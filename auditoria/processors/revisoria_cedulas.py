from openpyxl.worksheet.worksheet import Worksheet
from audits.models import Audit
from auditoria.models import RevisoriaSubcuenta
from auditoria.revisoria.common import (
    _extraer_fechas_desde_db,
    _formatear_fecha,
)


def llenar_cedula_31_efectivo_y_equivalentes(ws: Worksheet, audit: Audit):
    """
    Llena la hoja A-SUM (3.1 EFECTIVO Y EQUIVALENTES) con las subcuentas
    guardadas en RevisoriaSubcuenta para 'Efectivo y Equivalentes'.
    """

    # ---------- 0️⃣ Fechas de encabezado ----------
    fechas = _extraer_fechas_desde_db(audit.id)
    d1 = fechas.get("d1")
    d2 = fechas.get("d2")
    d3 = fechas.get("d3")
    d4 = fechas.get("d4")

    fila_fechas = 13  # donde están los encabezados tipo "Al 01/01/2023"

    if d1:
        ws.cell(row=fila_fechas, column=4, value=f"Al {_formatear_fecha(d1)}")
    if d2:
        ws.cell(row=fila_fechas, column=5, value=f"Al {_formatear_fecha(d2)}")
    if d3:
        ws.cell(row=fila_fechas, column=6, value=f"Al {_formatear_fecha(d3)}")
    if d4:
        ws.cell(row=fila_fechas, column=9, value=f"Al {_formatear_fecha(d4)}")

    # ---------- 1️⃣ Subcuentas desde BD ----------
    subcuentas = (
        RevisoriaSubcuenta.objects
        .filter(
            audit=audit,
            seccion__iexact="Activo",
            cuenta_principal__iexact="Efectivo y Equivalentes",
        )
        .order_by("nombre_subcuenta")
    )

    # Si no hay nada, no tocamos el bloque
    if not subcuentas.exists():
        return

    # ---------- 2️⃣ Ajustar dinámicamente las filas ----------
    fila_inicio = 14  # primera fila de datos de subcuenta
    fila_suma = None

    # Buscar la fila que contiene el texto "Suma" (en col A/B/C)
    for r in range(fila_inicio, ws.max_row + 1):
        for c in (1, 2, 3):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().lower() == "suma":
                fila_suma = r
                break
        if fila_suma:
            break

    if not fila_suma:
        # Si por alguna razón no encontramos "Suma", no tocamos filas
        fila_suma = fila_inicio + len(subcuentas) + 1

    filas_actuales = fila_suma - fila_inicio
    filas_necesarias = subcuentas.count()

    if filas_necesarias > filas_actuales:
        # Insertar filas adicionales antes de la fila "Suma"
        ws.insert_rows(fila_suma, filas_necesarias - filas_actuales)
        fila_suma += (filas_necesarias - filas_actuales)
    elif filas_necesarias < filas_actuales:
        # Limpiar filas sobrantes entre los datos y "Suma"
        for r in range(fila_inicio + filas_necesarias, fila_suma):
            for c in range(1, 10):  # columnas A..I
                ws.cell(row=r, column=c).value = None

    # ---------- 3️⃣ Escribir valores ----------
    row = fila_inicio

    for idx, s in enumerate(subcuentas, start=1):
        # Col A → Código (numérico incremental)
        ws.cell(row=row, column=1, value=idx)

        # Col B → Subcuenta
        ws.cell(row=row, column=2, value=s.nombre_subcuenta)

        # Col D/E/F → Saldos s/ balance general (tres cortes)
        ws.cell(row=row, column=4, value=s.saldo_ini_anterior)
        ws.cell(row=row, column=5, value=s.saldo_julio_anterior)
        ws.cell(row=row, column=6, value=s.saldo_dic_anterior)

        # Col G/H → Ajustes / Reclasificaciones
        ws.cell(row=row, column=7, value=s.ajuste_debe)
        ws.cell(row=row, column=8, value=s.ajuste_haber)

        # Col I → Saldos según auditoría (saldo dic actual)
        ws.cell(row=row, column=9, value=s.saldo_dic_actual)

        row += 1

def llenar_cedula_33_cuentas_por_cobrar(ws: Worksheet, audit: Audit):
    """
    Llena la hoja 'C CUENTAS POR COBRAR' con las subcuentas
    guardadas en RevisoriaSubcuenta para 'Cuentas por Cobrar'.
    """

    # ---------- 0️⃣ Fechas de encabezado ----------
    # Usamos el mismo helper que para la centralizadora
    fechas = _extraer_fechas_desde_db(audit.id)
    d1 = fechas.get("d1")  # corte 1
    d2 = fechas.get("d2")  # corte 2
    d3 = fechas.get("d3")  # corte 3
    d4 = fechas.get("d4")  # año actual

    # En esta cédula, la fila de fechas es la 11 (debajo de los títulos y encima de los datos)
    fila_fechas = 11

    if d1:
        ws.cell(row=fila_fechas, column=4, value=f"Al {_formatear_fecha(d1)}")  # Col D
    if d2:
        ws.cell(row=fila_fechas, column=5, value=f"Al {_formatear_fecha(d2)}")  # Col E
    if d3:
        ws.cell(row=fila_fechas, column=6, value=f"Al {_formatear_fecha(d3)}")  # Col F
    if d4:
        ws.cell(row=fila_fechas, column=9, value=f"Al {_formatear_fecha(d4)}")  # Col I

    # ---------- 1️⃣ Subcuentas desde BD ----------
    subcuentas = (
        RevisoriaSubcuenta.objects
        .filter(
            audit=audit,
            seccion__iexact="Activo",
            cuenta_principal__iexact="Cuentas por Cobrar",
        )
        .order_by("nombre_subcuenta")
    )

    if not subcuentas.exists():
        return

    # ---------- 2️⃣ Ajustar dinámicamente las filas ----------
    # Los datos empiezan en la fila 12 (debajo de la fila de fechas)
    fila_inicio = 12
    fila_suma = None

    # Buscar la fila donde está la palabra "Suma" (en A/B/C)
    for r in range(fila_inicio, ws.max_row + 1):
        for c in (1, 2, 3):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().lower() == "suma":
                fila_suma = r
                break
        if fila_suma:
            break

    if not fila_suma:
        # fallback por si alguien cambió la plantilla
        fila_suma = fila_inicio + subcuentas.count() + 1

    filas_actuales = fila_suma - fila_inicio
    filas_necesarias = subcuentas.count()

    if filas_necesarias > filas_actuales:
        # Insertar filas antes de "Suma"
        ws.insert_rows(fila_suma, filas_necesarias - filas_actuales)
        fila_suma += (filas_necesarias - filas_actuales)
    elif filas_necesarias < filas_actuales:
        # Limpiar filas sobrantes entre datos y "Suma"
        for r in range(fila_inicio + filas_necesarias, fila_suma):
            for c in range(1, 10):  # columnas A..I
                ws.cell(row=r, column=c).value = None

    # ---------- 3️⃣ Escribir valores ----------
    row = fila_inicio

    for idx, s in enumerate(subcuentas, start=1):
        # Col A → Código consecutivo
        ws.cell(row=row, column=1, value=idx)

        # Col B → Subcuenta
        ws.cell(row=row, column=2, value=s.nombre_subcuenta)

        # Col D/E/F → Saldos s/ balance general
        ws.cell(row=row, column=4, value=s.saldo_ini_anterior)
        ws.cell(row=row, column=5, value=s.saldo_julio_anterior)
        ws.cell(row=row, column=6, value=s.saldo_dic_anterior)

        # Col G/H → Ajustes / Reclasificaciones
        ws.cell(row=row, column=7, value=s.ajuste_debe)
        ws.cell(row=row, column=8, value=s.ajuste_haber)

        # Col I → Saldos según auditoría
        ws.cell(row=row, column=9, value=s.saldo_dic_actual)

        row += 1

def llenar_cedula_34_inventarios(ws: Worksheet, audit: Audit):
    """
    Llena la hoja 'D-SUM' (3.4 INVENTARIOS) con las subcuentas
    guardadas en RevisoriaSubcuenta para 'Inventarios'.

    Mapea así:
      - Col A (Código): 1, 2, 3...
      - Col B (Subcuenta): nombre de la subcuenta
      - Col D: saldo_ini_anterior      (Al 01/01/AAAA)
      - Col E: saldo_julio_anterior    (Al 31/07/AAAA)
      - Col F: saldo_dic_anterior      (Al 31/12 año anterior)
      - Col G: ajuste_debe
      - Col H: ajuste_haber
      - Col I: saldo_dic_actual        (Al 31/12 año actual / “Saldos según auditoría”)
    """

    # ---------- 0️⃣ Fechas de encabezado ----------
    # Usa las mismas fechas que el balance (d1..d4)
    fechas = _extraer_fechas_desde_db(audit.id)
    d1 = fechas.get("d1")
    d2 = fechas.get("d2")
    d3 = fechas.get("d3")
    d4 = fechas.get("d4")

    # En esta cédula, las fechas van en la fila 12 (debajo de los encabezados)
    fila_fechas = 12

    if d1:
        ws.cell(row=fila_fechas, column=4, value=f"Al {_formatear_fecha(d1)}")
    if d2:
        ws.cell(row=fila_fechas, column=5, value=f"Al {_formatear_fecha(d2)}")
    if d3:
        ws.cell(row=fila_fechas, column=6, value=f"Al {_formatear_fecha(d3)}")
    if d4:
        ws.cell(row=fila_fechas, column=9, value=f"Al {_formatear_fecha(d4)}")

    # ---------- 1️⃣ Subcuentas desde BD ----------
    subcuentas = (
        RevisoriaSubcuenta.objects
        .filter(
            audit=audit,
            seccion__iexact="Activo",
            cuenta_principal__iexact="Inventarios",
        )
        .order_by("nombre_subcuenta")
    )

    if not subcuentas.exists():
        # Si no hay subcuentas para esta cuenta, no tocamos la cédula
        return

    # ---------- 2️⃣ Ajustar dinámicamente las filas ----------
    # En esta cédula, los datos empiezan en la fila 13
    fila_inicio = 13
    fila_suma = None

    # Buscar la fila donde está la palabra "Suma" (en A/B/C)
    for r in range(fila_inicio, ws.max_row + 1):
        for c in (1, 2, 3):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().lower() == "suma":
                fila_suma = r
                break
        if fila_suma:
            break

    if not fila_suma:
        # fallback por si alguien cambió la plantilla
        fila_suma = fila_inicio + subcuentas.count() + 1

    filas_actuales = fila_suma - fila_inicio
    filas_necesarias = subcuentas.count()

    if filas_necesarias > filas_actuales:
        # Insertar filas antes de "Suma"
        ws.insert_rows(fila_suma, filas_necesarias - filas_actuales)
        fila_suma += (filas_necesarias - filas_actuales)
    elif filas_necesarias < filas_actuales:
        # Limpiar filas sobrantes entre datos y "Suma"
        for r in range(fila_inicio + filas_necesarias, fila_suma):
            for c in range(1, 10):  # columnas A..I
                ws.cell(row=r, column=c).value = None

    # ---------- 3️⃣ Escribir valores ----------
    row = fila_inicio

    for idx, s in enumerate(subcuentas, start=1):
        # Col A → Código consecutivo
        ws.cell(row=row, column=1, value=idx)

        # Col B → Subcuenta
        ws.cell(row=row, column=2, value=s.nombre_subcuenta)

        # Col D/E/F → Saldos s/ balance general
        ws.cell(row=row, column=4, value=s.saldo_ini_anterior)
        ws.cell(row=row, column=5, value=s.saldo_julio_anterior)
        ws.cell(row=row, column=6, value=s.saldo_dic_anterior)

        # Col G/H → Ajustes / Reclasificaciones
        ws.cell(row=row, column=7, value=s.ajuste_debe)
        ws.cell(row=row, column=8, value=s.ajuste_haber)

        # Col I → Saldos según auditoría
        ws.cell(row=row, column=9, value=s.saldo_dic_actual)

        row += 1
        
