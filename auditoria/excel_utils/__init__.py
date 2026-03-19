"""
Utilidades modularizadas para procesamiento de documentos Excel.
Funciones principales para modificar documentos Excel normales y con macros.
"""

from openpyxl import load_workbook
from .date_formatter import format_audit_dates
from .xlsm_processor import modify_document_excel_with_macros
from ..processors.excel.sheet_processor import process_excel_sheets
from ..utils.replacements_utils import (
    get_replacements_config,
    get_tables_config,
    build_replacements_dict,
)
from ..utils.data_db import get_all_financial_data

from auditoria.revisoria.otros_pasivo_patrimonio import (
    update_cedula_411_otras,
    update_cedula_421_otras,
)
from pathlib import Path
from auditoria.revisoria.Estado_resul import (
    update_cedula_ingresos_subcuentas,
    update_cedula_costos_gastos_subcuentas,
)


def modify_document_excel(template_path, audit):
    wb = load_workbook(template_path)

    # 1️⃣ Formatear fechas
    fecha_inicio, fecha_fin = format_audit_dates(audit)

    # 2️⃣ Obtener datos financieros de BD
    financial_data = get_all_financial_data(audit.id)
    data_bd = financial_data["organized"]

    # 3️⃣ Configs de reemplazos estáticos
    replacements_config = get_replacements_config()
    tables_config = get_tables_config()

    # 4️⃣ Actualizar bloque dinámico del Balance (centralizadoras 2, 3, 4 y 5)
    try:

        # 🔍 Nombre del archivo (para distinguir 2, 3, 4, 5, 5.1, etc.)
        file_name_upper = str(template_path).upper()

        # 🔍 Detección robusta de hoja financiera
        sheet_financiero = None
        pistas = ["ESTADOS FINANCIEROS", "ACTIVO", "Fecha corte año actual"]

        for name in wb.sheetnames:
            sh = wb[name]
            encontrado = False

            for row in sh.iter_rows(values_only=True):
                row_texto = " ".join(str(v) for v in row if v is not None)
                if any(pista.lower() in row_texto.lower() for pista in pistas):
                    encontrado = True
                    break

            if encontrado:
                sheet_financiero = sh
                break

        # 👇 Solo aplicamos el balance dinámico genérico si el archivo tiene una hoja 'Hoja1'
        if "Hoja1" in wb.sheetnames and getattr(audit, 'activa_revisoria', False):
            from auditoria.revisoria.activo import update_dynamic_excel  # import local para no tocar otros lados

            if sheet_financiero:
                wb = update_dynamic_excel(
                    workbook=wb,
                    sheet_name="Hoja1",
                    audit_id=audit.id,
                    sheet_financiero=sheet_financiero,
                )
            else:
                wb = update_dynamic_excel(
                    workbook=wb,
                    sheet_name="Hoja1",
                    audit_id=audit.id,
                    sheet_financiero=None,
                )
        else:
            # 🔹 Caso de las cédulas 3.x, 4.x, 5.x y demás archivos sin 'Hoja1'
            # print("ℹ️ Este archivo no tiene hoja 'Hoja1'; se omite actualización dinámica del balance.")
            pass

        # 🟣 EXTRA 1: si el archivo es la 5 CENTRALIZADORA DEL ESTADO DE RESULTADOS,
        # actualizamos también su bloque propio con las cuentas de ESTADO DE RESULTADOS.
        if "5 CENTRALIZADORA" in file_name_upper and "ESTADO DE RESULTADOS" in file_name_upper:
            from auditoria.revisoria.Estado_resul import (
                update_centralizadora_estado_resultados,
            )

            sheet_name_er = "Hoja1"  # normalmente la centralizadora usa 'Hoja1'
            if sheet_name_er in wb.sheetnames:
                wb = update_centralizadora_estado_resultados(
                    workbook=wb,
                    sheet_name=sheet_name_er,
                    audit_id=audit.id,
                    sheet_financiero=None,
                )
                print("✔️ CENTRALIZADORA DEL ESTADO DE RESULTADOS actualizada ✔️")
            else:
                print("⚠️ No se encontró la hoja 'Hoja1' en la CENTRALIZADORA DEL ESTADO DE RESULTADOS.")

    except Exception as e:
        print("❌ Error en actualización dinámica:", e)

    # 4️⃣.bis Cédulas de Revisoria (3.1, 3.3, 3.4) basadas en RevisoriaSubcuenta
    try:
        from pathlib import Path
        from auditoria.revisoria.activo import (
            update_cedula_31_efectivo,
            update_cedula_33_cuentas_por_cobrar,
            update_cedula_34_inventarios,
        )

        file_name = Path(template_path).name.upper()

        # 🔹 3.1 EFECTIVO Y EQUIVALENTES
        es_cedula_31 = (
            "3.1 EFECTIVO_Y_EQUIVALENTES" in file_name
            or "3.1 EFECTIVO Y EQUIVALENTES" in file_name
        )

        # 🔹 3.3 CUENTAS POR COBRAR
        es_cedula_33 = (
            "3.3 CUENTAS_POR_COBRAR" in file_name
            or "3.3 CUENTAS POR COBRAR" in file_name
        )

        # 🔹 3.4 INVENTARIOS
        es_cedula_34 = "3.4 INVENTARIOS" in file_name

        if es_cedula_31:

            if "A-SUM" in wb.sheetnames:
                wb = update_cedula_31_efectivo(
                    workbook=wb,
                    sheet_name="A-SUM",
                    audit_id=audit.id,
                    sheet_financiero=None,  # no se usa aquí
                )
            else:
                print("⚠️ Hoja 'A-SUM' no encontrada en esta plantilla; se omite cédula 3.1.")

        elif es_cedula_33:
            # Nombre de la hoja según la plantilla que mostraste
            sheet_name_33 = "C CUENTAS POR COBRAR"

            if sheet_name_33 in wb.sheetnames:
                wb = update_cedula_33_cuentas_por_cobrar(
                    workbook=wb,
                    sheet_name=sheet_name_33,
                    audit_id=audit.id,
                    sheet_financiero=None,  # igual que en 3.1, no se usa
                )
            else:
                print(f"⚠️ Hoja '{sheet_name_33}' no encontrada en esta plantilla; se omite cédula 3.3.")

        elif es_cedula_34:

            sheet_name_34 = "D-SUM"

            if sheet_name_34 in wb.sheetnames:
                wb = update_cedula_34_inventarios(
                    workbook=wb,
                    sheet_name=sheet_name_34,
                    audit_id=audit.id,
                    sheet_financiero=None,  # no se usa aquí
                )
            else:
                print(f"⚠️ Hoja '{sheet_name_34}' no encontrada en esta plantilla; se omite cédula 3.4.")

    except Exception as e:
        print("⚠️ Error al actualizar cédulas de Revisoria (3.1 / 3.3 / 3.4):", e)

    # 4️⃣.ter Cédulas 3.6 OTRAS CÉDULAS SUMARIAS (bloque independiente)
    try:
        from pathlib import Path
        from auditoria.revisoria.activo import (
            update_cedula_36_otras,
        )  # 👈 asegúrate que exista con EXACTO este nombre

        file_name = Path(template_path).name.upper()

        es_cedula_36 = (
            "3.6 OTRAS CEDULAS SUMARIAS" in file_name
            or "3.6 OTRAS CÉDULAS SUMARIAS" in file_name
            or "3.6 OTRAS" in file_name
        )

        if es_cedula_36:

            wb = update_cedula_36_otras(
                workbook=wb,
                audit_id=audit.id,
            )

    except Exception as e:
        print("⚠️ Error al actualizar cédulas 3.6 OTRAS CÉDULAS SUMARIAS:", e)

    # 4️⃣.quater Cédulas 4.1.1 y 4.2.1 OTRAS CÉDULAS SUMARIAS (PASIVO / PATRIMONIO)
    try:
        from pathlib import Path

        file_name = Path(template_path).name.upper()

        # 🔹 4.1.1 OTRAS CÉDULAS SUMARIAS DE PASIVO
        es_cedula_411 = "4.1.1" in file_name

        # 🔹 4.2.1 OTRAS CÉDULAS SUMARIAS DE PATRIMONIO
        es_cedula_421 = "4.2.1" in file_name

        if es_cedula_411:

            wb = update_cedula_411_otras(
                workbook=wb,
                audit_id=audit.id,
                sheet_financiero=None,
            )

        if es_cedula_421:
            print("🔹 Detectada plantilla de cédula 4.2.1 OTRAS CÉDULAS SUMARIAS DE PATRIMONIO…")

            wb = update_cedula_421_otras(
                workbook=wb,
                audit_id=audit.id,
                sheet_financiero=None,
            )

    except Exception as e:
        print("⚠️ Error al actualizar cédulas 4.1.1 / 4.2.1 OTRAS CÉDULAS SUMARIAS:", e)

    # 4️⃣.quinquies Cédulas 5.1 y 5.2 (INGRESOS y COSTOS Y GASTOS)
    try:

        file_name = Path(template_path).name.upper()

        es_cedula_51 = "5.1 INGRESOS" in file_name
        es_cedula_52 = (
            "5.2 COSTOS_Y_GASTOS" in file_name
            or "5.2 COSTOS Y GASTOS" in file_name
        )

        if es_cedula_51:
            if "H-SUM" in wb.sheetnames:
                wb = update_cedula_ingresos_subcuentas(
                    workbook=wb,
                    audit_id=audit.id,
                    sheet_name="H-SUM",
                )
            else:
                print("⚠️ [INGRESOS] Hoja 'H-SUM' no encontrada en esta plantilla.")

        if es_cedula_52:

            if "I-SUM" in wb.sheetnames:
                wb = update_cedula_costos_gastos_subcuentas(
                    workbook=wb,
                    audit_id=audit.id,
                    sheet_name="I-SUM",
                )
            else:
                print("⚠️ [COSTOS/GASTOS] Hoja 'I-SUM' no encontrada en esta plantilla.")

    except Exception as e:
        print("⚠️ Error al actualizar cédulas 5.x de Estado de Resultados:", e)

    # 4️⃣.sexies Cédula 5.4 OTRAS CÉDULAS SUMARIAS (Estado de Resultados)
    try:
        file_name_upper = str(template_path).upper()

        if "5.4" in file_name_upper and "OTRAS CEDULAS SUMARIAS" in file_name_upper:
            from auditoria.revisoria.Estado_resul import (
                update_cedula_54_otras_estado_resultados,
            )


            wb = update_cedula_54_otras_estado_resultados(
                workbook=wb,
                audit_id=audit.id,
                sheet_financiero=None,
            )

    except Exception as e:
        print("⚠️ Error al actualizar cédulas 5.4 OTRAS CÉDULAS SUMARIAS:", e)

    # 5️⃣ Reemplazos de texto en celdas / tablas
    # Obtener los reemplazos básicos
    replacements = build_replacements_dict(
        config=replacements_config,
        audit=audit,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    # Procesar el documento Excel
    process_excel_sheets(wb, tables_config, replacements, data_bd, template_path)

    return wb

# Exportar las funciones principales para compatibilidad hacia atrás
__all__ = [
    'modify_document_excel',
    'modify_document_excel_with_macros',
    'format_audit_dates'
]
