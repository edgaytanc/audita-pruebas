# auditoria/Estado_resul.py

from datetime import date

from auditoria.models import BalanceCuentas
from .common import (
    _clean,
    _normalizar,
    _set_cell_value,
    _formatear_fecha,
    _extraer_fechas_desde_db,
)

from openpyxl import load_workbook
from audits.models import Audit
from auditoria.models import RevisoriaSubcuenta



# ================== BUSCAR FILAS CLAVE ==================

def _encontrar_fila_codigo_resultados(sheet):
    """
    Busca la fila donde aparece el encabezado 'Sub Cuenta' / 'Subcuenta'
    (columna B o C). Esa será la fila de encabezado del bloque
    de detalle del ESTADO DE RESULTADOS.
    """
    for row in range(1, sheet.max_row + 1):
        for col in (2, 3):
            valor = sheet.cell(row=row, column=col).value
            if "subcuenta" in _clean(valor):
                return row
    return None


def _encontrar_fila_suma_resultados(sheet, fila_inicio):
    """
    Busca la fila donde está el texto 'Suma' en las columnas B o C,
    a partir de fila_inicio. Esa fila se considera el cierre
    del bloque de detalle de ESTADO DE RESULTADOS.
    """
    if not fila_inicio:
        return None

    for row in range(fila_inicio + 1, sheet.max_row + 1):
        for col in (2, 3):
            valor = sheet.cell(row=row, column=col).value
            if "suma" in _clean(valor):
                return row
    return None


# ================== OBTENER DATOS DB (ESTADO RESULTADOS) ==================

def _obtener_cuentas_resultados(audit_id):
    """
    Arma una estructura por cuenta con d1, d2, d3, d4
    usando BalanceCuentas (tipo_balance = ANUAL, seccion ≈ Estado de Resultados).

    Usa las mismas fechas que `_extraer_fechas_desde_db` (d1..d4)
    para que quede alineado con activo/pasivo/patrimonio.
    """

    # 1) Fechas d1..d4 desde BD
    fechas = _extraer_fechas_desde_db(audit_id)

    d1 = fechas.get("d1")
    d2 = fechas.get("d2")
    d3 = fechas.get("d3")
    d4 = fechas.get("d4")

    # 🔹 Filtro más flexible: cualquier seccion que contenga "RESULTADO"
    qs = (
        BalanceCuentas.objects
        .filter(
            audit_id=audit_id,
            tipo_balance="ANUAL",
            seccion__icontains="RESULTADO",
        )
        .order_by("nombre_cuenta", "fecha_corte")
    )

    cuentas = {}

    for bc in qs:
        nombre = (bc.nombre_cuenta or "").strip()
        if not nombre:
            continue

        f = bc.fecha_corte
        slot = None

        # Mapeo 1:1 contra las fechas clasificadas
        tc = (bc.tipo_cuenta or "").strip().upper()
        if d1 and f == d1:
            slot = "d1"
        elif d2 and f == d2:
            slot = "d2"
        # ✅ Si d3 y d4 son la MISMA fecha, distinguir por tipo_cuenta
        elif d3 and d4 and d3 == d4 and f == d3:
            slot = "d4" if tc == "NT_ACT" else "d3"
        elif d3 and f == d3:
            slot = "d3"
        elif d4 and f == d4:
            slot = "d4"



        if not slot:
            # Fecha que no coincide con ninguno de los cortes oficiales
            continue

        if nombre not in cuentas:
            cuentas[nombre] = {
                "nombre": nombre,
                "d1": None,
                "d2": None,
                "d3": None,
                "d4": None,
            }

        cuentas[nombre][slot] = float(bc.valor)


    # Resumen por columna
    totales_por_slot = {"d1": 0, "d2": 0, "d3": 0, "d4": 0}
    for c in cuentas.values():
        for slot in totales_por_slot.keys():
            if c[slot] is not None:
                totales_por_slot[slot] += 1

    return list(cuentas.values())


def _obtener_ajustes_resultados(audit_id):
    """
    Devuelve un dict con Debe/Haber por cuenta para los AJUSTES
    del Estado de Resultados.

    {
        'Ingresos': {'debe': 150000.0, 'haber': 120000.0},
        ...
    }
    """
    qs = BalanceCuentas.objects.filter(
        audit_id=audit_id,
        tipo_balance="AJUSTE",
        seccion__icontains="RESULTADO",  # 🔹 también flexible
    )

    ajustes = {}
    for bc in qs:
        nombre = (bc.nombre_cuenta or "").strip()
        if not nombre:
            continue

        if nombre not in ajustes:
            ajustes[nombre] = {"debe": None, "haber": None}

        tipo = (bc.tipo_cuenta or "").upper()
        if tipo == "DEBE":
            ajustes[nombre]["debe"] = float(bc.valor)
        elif tipo == "HABER":
            ajustes[nombre]["haber"] = float(bc.valor)

    return ajustes


def _obtener_cuentas_y_ajustes_resultados(audit_id):
    """
    Helper que devuelve:
      - cuentas: lista de dicts {nombre, d1, d2, d3, d4}
      - ajustes: dict { nombre_cuenta: {"debe": float, "haber": float} }
    """
    cuentas = _obtener_cuentas_resultados(audit_id)
    ajustes = _obtener_ajustes_resultados(audit_id)
    return cuentas, ajustes


# ================== ESCRIBIR EN EXCEL (CENTRALIZADORA RESULTADOS) ==================

def _actualizar_bloque_estado_resultados(sheet, audit_id, cuentas=None, ajustes=None):
    """
    Actualiza el bloque de detalle en '5 CENTRALIZADORA  DEL ESTADO DE RESULTADOS.xlsx'.
    """

    # Buscar encabezado
    fila_header = _encontrar_fila_codigo_resultados(sheet)

    # Fallback si no encuentra el encabezado: asumimos fila 13 (igual que ACTIVO)
    if fila_header is None:
        print("ℹ [RESULTADOS] No se encontró 'Subcuenta'; usando fallback fila 13")
        fila_header = 13

    fila_suma = _encontrar_fila_suma_resultados(sheet, fila_header)

    # Si no llegaron cuentas/ajustes, los obtenemos de BD
    if cuentas is None or ajustes is None:
        cuentas, ajustes = _obtener_cuentas_y_ajustes_resultados(audit_id)

    if not fila_suma or not cuentas:
        print("⚠ No se pudo localizar bloque ESTADO DE RESULTADOS (suma/cuentas faltan)")
        return

    fila_inicio = fila_header + 1
    filas_actuales = fila_suma - fila_inicio
    filas_necesarias = len(cuentas)

    # Ajustar filas dinámicamente (igual que en ACTIVO)
    if filas_necesarias > filas_actuales:
        # Insertar filas adicionales antes de la fila de sumas
        sheet.insert_rows(fila_suma, filas_necesarias - filas_actuales)
    elif filas_necesarias < filas_actuales:
        # Limpiar filas sobrantes entre detalle y sumas
        for r in range(fila_inicio + filas_necesarias, fila_suma):
            for c in range(1, 20):
                sheet.cell(row=r, column=c).value = None

    # Columnas concretas del template
    COL_COD   = 2  # B → Código
    COL_NOM   = 3  # C → Sub Cuenta
    COL_REF   = 4  # D → Ref P/T
    COL_D1    = 5  # E → Saldo 1
    COL_D2    = 6  # F → Saldo 2
    COL_D3    = 7  # G → Saldo 3
    COL_DEBE  = 8  # H → Debe (ajustes)
    COL_HABER = 9  # I → Haber (ajustes)
    COL_D4    = 10 # J → Saldo 4 (año actual)

    for i, cta in enumerate(cuentas):
        r = fila_inicio + i
        nombre = cta["nombre"]

        # Valores base
        v_d1 = cta.get("d1")
        v_d2 = cta.get("d2")
        v_d3 = cta.get("d3")
        v_d4 = cta.get("d4")

        # Parche igual que en ACTIVO: si d1 viene vacío pero d4 tiene valor,
        # usamos d4 también en la 1ª columna de saldos.
        if v_d1 is None and v_d4 is not None:
            v_d1 = v_d4

        datos_ajuste = ajustes.get(nombre, {}) if ajustes else {}
        v_debe = datos_ajuste.get("debe")
        v_haber = datos_ajuste.get("haber")

        # Escribir base
        sheet.cell(row=r, column=COL_COD).value = i + 1
        sheet.cell(row=r, column=COL_NOM).value = nombre
        sheet.cell(row=r, column=COL_REF).value = ""  # Ref P/T vacío
        sheet.cell(row=r, column=COL_D1).value = v_d1
        sheet.cell(row=r, column=COL_D2).value = v_d2
        sheet.cell(row=r, column=COL_D3).value = v_d3
        sheet.cell(row=r, column=COL_D4).value = v_d4

        # Debe / Haber desde ajustes (si existen)
        if datos_ajuste:
            sheet.cell(row=r, column=COL_DEBE).value = v_debe
            sheet.cell(row=r, column=COL_HABER).value = v_haber

    print("✔️ ESTADO DE RESULTADOS → Base actualizada (CENTRALIZADORA)")


# ================== FLUJO PRINCIPAL PÚBLICO ==================

def update_centralizadora_estado_resultados(workbook, sheet_name, audit_id, sheet_financiero=None):
    """
    Actualiza el archivo '5 CENTRALIZADORA  DEL ESTADO DE RESULTADOS.xlsx'.

    - Rellena el bloque de detalle (Sub Cuenta) con d1..d4 según BalanceCuentas
      (seccion≈Estado de Resultados, tipo_balance='ANUAL').
    - Aplica Debe/Haber si existen registros de AJUSTE para cada cuenta.
    - NO toca fórmulas ni encabezados de la plantilla;
      solo rellena la zona de detalle bajo 'Sub Cuenta'.
    """
    sheet = workbook[sheet_name]

    # 1️⃣ Traer cuentas + ajustes solo de ESTADO DE RESULTADOS
    cuentas, ajustes = _obtener_cuentas_y_ajustes_resultados(audit_id)

    if not cuentas:
        print("⚠ No hay cuentas de ESTADO DE RESULTADOS en BD para esta auditoría.")
        return workbook

    # 2️⃣ Escribir bloque de detalle
    _actualizar_bloque_estado_resultados(sheet, audit_id, cuentas=cuentas, ajustes=ajustes)

    print("✔️ CENTRALIZADORA DEL ESTADO DE RESULTADOS actualizada ✔️")
    return workbook

def _encontrar_fila_detalle_ingresos(sheet):
    """
    Busca la fila de encabezado de detalle en la hoja H-SUM
    (por ejemplo donde aparece 'Sub Cuenta', 'Subcuenta' o 'Detalle'),
    y devuelve la fila siguiente (donde empiezan las subcuentas).
    """
    for row in range(1, sheet.max_row + 1):
        for col in (1, 2, 3):  # A, B, C
            valor = sheet.cell(row=row, column=col).value
            txt = _clean(valor)
            compacto = txt.replace(" ", "")

            if (
                "subcuenta" in compacto               # 'Sub Cuenta' / 'Subcuenta'
                or ("sub" in txt and "cuenta" in txt) # más laxo
                or "detalle" in txt                   # por si dice 'Detalle'
            ):
                # La fila de detalle empieza una fila debajo del encabezado
                return row + 1
    return None


def _encontrar_fila_total_ingresos(sheet, fila_inicio_detalle):
    """
    Busca la fila donde aparece el TOTAL / SUMA de la cédula de ingresos.
    Se detiene cuando encuentra una fila con 'total' o 'suma'.
    """
    if not fila_inicio_detalle:
        return None

    for row in range(fila_inicio_detalle, sheet.max_row + 1):
        for col in (1, 2, 3):  # A, B, C
            valor = sheet.cell(row=row, column=col).value
            txt = _clean(valor)
            if "total" in txt or "suma" in txt:  # aquí capturamos el 'Suma' de tu plantilla
                return row
    return None


def update_cedula_ingresos_subcuentas(workbook, audit_id, sheet_name="H-SUM"):
    """
    Rellena la hoja H-SUM del archivo '5.1 INGRESOS.xlsx' con las
    subcuentas de INGRESOS (Estado de Resultados).

    Origen de datos:
      - RevisoriaSubcuenta (seccion='EstadoResultado', cuenta_principal='Ingresos')
    """
    if sheet_name not in workbook.sheetnames:
        print(f"⚠️ [INGRESOS] Hoja '{sheet_name}' no encontrada en el workbook.")
        return workbook

    sheet = workbook[sheet_name]
    print(f"🔄 [INGRESOS] Actualizando cédula de subcuentas en hoja: {sheet.title}")

    # 0️⃣ Fechas d1..d4 desde BD (para el encabezado de columnas)
    fechas = _extraer_fechas_desde_db(audit_id)

    d1 = fechas.get("d1")
    d2 = fechas.get("d2")
    d3 = fechas.get("d3")
    d4 = fechas.get("d4")

    # Localizar fila donde está el encabezado "SALDOS S/ BALANCE GENERAL"
    fila_enc_saldos = None
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            valor = sheet.cell(row=row, column=col).value
            txt = _clean(valor)
            if "saldos" in txt and "balance" in txt:
                fila_enc_saldos = row
                break
        if fila_enc_saldos:
            break

    if fila_enc_saldos:
        fila_fechas = fila_enc_saldos + 1  # igual patrón que CENTRALIZADORA

        # E,F,G,J → d1,d2,d3,d4
        if d1:
            sheet.cell(row=fila_fechas, column=5).value = _formatear_fecha(d1)
            print(f"✔ Fecha d1 aplicada en H-SUM ({fila_fechas}, 5) → {_formatear_fecha(d1)}")
        if d2:
            sheet.cell(row=fila_fechas, column=6).value = _formatear_fecha(d2)
            print(f"✔ Fecha d2 aplicada en H-SUM ({fila_fechas}, 6) → {_formatear_fecha(d2)}")
        if d3:
            sheet.cell(row=fila_fechas, column=7).value = _formatear_fecha(d3)
            print(f"✔ Fecha d3 aplicada en H-SUM ({fila_fechas}, 7) → {_formatear_fecha(d3)}")
        if d4:
            sheet.cell(row=fila_fechas, column=10).value = _formatear_fecha(d4)
            print(f"✔ Fecha d4 aplicada en H-SUM ({fila_fechas}, 10) → {_formatear_fecha(d4)}")
    else:
        print("ℹ️ [INGRESOS] No se encontró encabezado de 'SALDOS S/ BALANCE GENERAL' para aplicar fechas.")

    # 1️⃣ Traer subcuentas de INGRESOS desde BD
    subcuentas_qs = RevisoriaSubcuenta.objects.filter(
        audit_id=audit_id,
        seccion="EstadoResultado",
        cuenta_principal__iexact="Ingresos",
    ).order_by("nombre_subcuenta")

    subcuentas = list(subcuentas_qs)

    if not subcuentas:
        print("ℹ️ [INGRESOS] No hay subcuentas registradas para 'Ingresos'.")
        return workbook


    # 2️⃣ Localizar bloque de detalle en H-SUM
    fila_detalle_inicio = _encontrar_fila_detalle_ingresos(sheet)
    fila_total = _encontrar_fila_total_ingresos(sheet, fila_detalle_inicio)

    if not fila_detalle_inicio or not fila_total:
        print("⚠️ [INGRESOS] No se pudo localizar correctamente el bloque de detalle (encabezado/total).")
        return workbook

    filas_actuales = fila_total - fila_detalle_inicio
    filas_necesarias = len(subcuentas)

    # 3️⃣ Ajustar filas dinámicamente
    if filas_necesarias > filas_actuales:
        sheet.insert_rows(fila_total, filas_necesarias - filas_actuales)

    elif filas_necesarias < filas_actuales:
        for r in range(fila_detalle_inicio + filas_necesarias, fila_total):
            for c in range(2, 11):  # B..J
                sheet.cell(row=r, column=c).value = None

    # Columnas de la cédula (B..J)
    COL_COD     = 2  # B -> Código
    COL_NOMBRE  = 3  # C -> Sub Cuenta
    COL_REF     = 4  # D -> Ref P/T
    COL_D1      = 5  # E -> saldo_ini_anterior
    COL_D2      = 6  # F -> saldo_julio_anterior
    COL_D3      = 7  # G -> saldo_dic_anterior
    COL_DEBE    = 8  # H -> ajuste_debe
    COL_HABER   = 9  # I -> ajuste_haber
    COL_D4      = 10 # J -> saldo_dic_actual

    for i, sub in enumerate(subcuentas):
        r = fila_detalle_inicio + i

        nombre_subcuenta = sub.nombre_subcuenta

        v_d1    = sub.saldo_ini_anterior
        v_d2    = sub.saldo_julio_anterior
        v_d3    = sub.saldo_dic_anterior
        v_debe  = sub.ajuste_debe
        v_haber = sub.ajuste_haber
        v_d4    = sub.saldo_dic_actual

        # Si d1 viene vacío pero d4 tiene valor, usamos d4 también en la 1ª columna.
        if v_d1 is None and v_d4 is not None:

            v_d1 = v_d4


        # Código secuencial
        sheet.cell(row=r, column=COL_COD).value = i + 1

        # Nombre subcuenta
        sheet.cell(row=r, column=COL_NOMBRE).value = nombre_subcuenta

        # Ref P/T vacío
        sheet.cell(row=r, column=COL_REF).value = ""

        # Valores
        sheet.cell(row=r, column=COL_D1).value = v_d1
        sheet.cell(row=r, column=COL_D2).value = v_d2
        sheet.cell(row=r, column=COL_D3).value = v_d3
        sheet.cell(row=r, column=COL_D4).value = v_d4

        # Debe / Haber
        sheet.cell(row=r, column=COL_DEBE).value = v_debe
        sheet.cell(row=r, column=COL_HABER).value = v_haber

    print("✔️ [INGRESOS] Cédula de subcuentas actualizada en H-SUM ✔️")
    return workbook

def update_cedula_costos_gastos_subcuentas(workbook, audit_id, sheet_name="I-SUM"):
    """
    Rellena la hoja I-SUM del archivo '5.2 COSTOS_Y_GASTOS.xlsx' con las
    subcuentas de COSTOS Y GASTOS (Estado de Resultados).

    Origen de datos:
      - RevisoriaSubcuenta (seccion='EstadoResultado',
                            cuenta_principal='Costos y Gastos')

    Columnas en I-SUM (misma estructura que H-SUM):
      A -> Código (correlativo)
      B -> Sub Cuenta
      C -> Ref P/T
      D -> saldo_ini_anterior   (Al d1)   → d1
      E -> saldo_julio_anterior (Al d2)   → d2
      F -> saldo_dic_anterior   (Al d3)   → d3
      G -> ajuste_debe          (Debe)
      H -> ajuste_haber         (Haber)
      I -> saldo_dic_actual     (Al d4)   → d4
    """
    if sheet_name not in workbook.sheetnames:
        print(f"⚠️ [COSTOS] Hoja '{sheet_name}' no encontrada en el workbook.")
        return workbook

    sheet = workbook[sheet_name]
    print(f"🔄 [COSTOS] Actualizando cédula de subcuentas en hoja: {sheet.title}")

    # 0️⃣ Encabezados de fechas (fila 12: D, E, F, I)
    fechas = _extraer_fechas_desde_db(audit_id)

    d1 = fechas.get("d1")
    d2 = fechas.get("d2")
    d3 = fechas.get("d3")
    d4 = fechas.get("d4")

    def _fmt_fecha(fecha):
        if not fecha:
            return None
        # _formatear_fecha ya sabe manejar date/datetime
        return _formatear_fecha(fecha)

    fila_fechas = 12

    if d1:
        sheet[f"D{fila_fechas}"].value = f"Al {_fmt_fecha(d1)}"
    if d2:
        sheet[f"E{fila_fechas}"].value = f"Al {_fmt_fecha(d2)}"
    if d3:
        sheet[f"F{fila_fechas}"].value = f"Al {_fmt_fecha(d3)}"
    if d4:
        sheet[f"I{fila_fechas}"].value = f"Al {_fmt_fecha(d4)}"

    # 1️⃣ Traer subcuentas de COSTOS Y GASTOS desde BD
    subcuentas_qs = RevisoriaSubcuenta.objects.filter(
        audit_id=audit_id,
        seccion="EstadoResultado",
        cuenta_principal__iexact="Costos y Gastos",
    ).order_by("nombre_subcuenta")

    subcuentas = list(subcuentas_qs)

    if not subcuentas:
        print("ℹ️ [COSTOS] No hay subcuentas registradas para 'Costos y Gastos'.")
        return workbook

    # 2️⃣ Localizar bloque de detalle en I-SUM
    fila_detalle_inicio = _encontrar_fila_detalle_ingresos(sheet)
    fila_total = _encontrar_fila_total_ingresos(sheet, fila_detalle_inicio)

    if not fila_detalle_inicio or not fila_total:
        print("⚠️ [COSTOS] No se pudo localizar correctamente el bloque de detalle (encabezado/total).")
        return workbook

    filas_actuales = fila_total - fila_detalle_inicio
    filas_necesarias = len(subcuentas)

    # 3️⃣ Ajustar filas dinámicamente (igual que en INGRESOS)
    if filas_necesarias > filas_actuales:
        sheet.insert_rows(fila_total, filas_necesarias - filas_actuales)

    elif filas_necesarias < filas_actuales:
        for r in range(fila_detalle_inicio + filas_necesarias, fila_total):
            for c in range(1, 15):
                sheet.cell(row=r, column=c).value = None


    # Columnas de la cédula (igual que H-SUM)
    COL_COD    = 1  # A
    COL_NOMBRE = 2  # B
    COL_REF    = 3  # C
    COL_D1     = 4  # D
    COL_D2     = 5  # E
    COL_D3     = 6  # F
    COL_DEBE   = 7  # G
    COL_HABER  = 8  # H
    COL_D4     = 9  # I


    for i, sub in enumerate(subcuentas):
        r = fila_detalle_inicio + i

        nombre_subcuenta = sub.nombre_subcuenta

        v_d1    = sub.saldo_ini_anterior
        v_d2    = sub.saldo_julio_anterior
        v_d3    = sub.saldo_dic_anterior
        v_debe  = sub.ajuste_debe
        v_haber = sub.ajuste_haber
        v_d4    = sub.saldo_dic_actual

        # Mismo parche que en ACTIVO / INGRESOS:
        if v_d1 is None and v_d4 is not None:

            v_d1 = v_d4


        # Código correlativo
        sheet.cell(row=r, column=COL_COD).value = i + 1

        # Sub Cuenta
        sheet.cell(row=r, column=COL_NOMBRE).value = nombre_subcuenta

        # Ref P/T (en blanco por ahora)
        sheet.cell(row=r, column=COL_REF).value = ""

        # Valores
        sheet.cell(row=r, column=COL_D1).value = v_d1
        sheet.cell(row=r, column=COL_D2).value = v_d2
        sheet.cell(row=r, column=COL_D3).value = v_d3
        sheet.cell(row=r, column=COL_D4).value = v_d4

        # Debe / Haber
        sheet.cell(row=r, column=COL_DEBE).value = v_debe
        sheet.cell(row=r, column=COL_HABER).value = v_haber

    print("✔️ [COSTOS] Cédula de subcuentas actualizada en I-SUM ✔️")
    return workbook

def update_cedula_54_otras_estado_resultados(workbook, audit_id, sheet_financiero=None):
    """
    Rellena la plantilla 5.4 OTRAS CÉDULAS SUMARIAS para ESTADO DE RESULTADOS.

    - Solo secciones que contengan 'EstadoResultado'
    - Excluye las cuentas principales 'Ingresos' y 'Costos y Gastos'
      (que ya tienen sus propias cédulas 5.1 y 5.2).
    - Cada cuenta_principal "otra" va a una hoja: '1', '2', '3', ...
    """

    from auditoria.models import RevisoriaSubcuenta

    # ⬅️ POSICIONES DE LAS FECHAS (ajusta si tu 5.4 tiene otras)
    CELDAS_FECHAS_54 = {
        "d1": ("D", 13),
        "d2": ("E", 13),
        "d3": ("F", 13),
        "d4": ("I", 13),
    }

    # 🔹 rango de filas SOLO para subcuentas (la de totales no se toca)
    AREA_DETALLE_INICIO = 14  # primera fila de detalle
    AREA_DETALLE_FIN    = 28  # última fila de detalle (p.ej. 29 = totales)

    # Cuentas de EstadoResultados que NO van aquí porque ya tienen cédula 5.1 / 5.2
    CUENTAS_ESPECIALES_ER = {
        "Ingresos",
        "Costos y Gastos",
    }

    try:
        # Todas las subcuentas de EstadoResultado que sí van a 5.4 (otras cuentas)
        qs = (
            RevisoriaSubcuenta.objects
            .filter(audit_id=audit_id, seccion="EstadoResultado")
            .exclude(cuenta_principal__in=CUENTAS_ESPECIALES_ER)
            .order_by("cuenta_principal", "id")
        )


        if not qs.exists():
            print("ℹ️ [5.4] No hay subcuentas 'otras' de EstadoResultados; se omite 5.4.")
            return workbook

        # 🔹 Fechas (misma lógica que otras cédulas, pero posiciones propias)
        fechas = _extraer_fechas_desde_db(audit_id)

        def es_placeholder_x(nombre: str) -> bool:
            """
            Devuelve True si el nombre es algo tipo 'XXXX', 'XXXXXXXXX', etc.
            (solo X y espacios/guiones).
            """
            if not nombre:
                return False
            limpio = nombre.replace(" ", "").replace("-", "").upper()
            return len(limpio) > 0 and set(limpio) == {"X"}

        # Agrupar solo las filas que realmente son subcuentas de detalle
        cuentas = {}
        for sub in qs:
            cp = (sub.cuenta_principal or "").strip()
            nombre = (sub.nombre_subcuenta or "").strip()

            if not nombre:
                continue

            # Saltar cabeceras (nombre igual a la cuenta principal)
            if nombre.upper() == cp.upper():
                continue

            # Saltar totales / placeholders tipo XXXX / XXXXXXXX
            if es_placeholder_x(nombre):
                continue

            cuentas.setdefault(cp, []).append(sub)

        for cp, subs in cuentas.items():
            print(f"   - {cp}: {len(subs)} subcuentas")

        # Rellenar las hojas 1, 2, 3,... con esas subcuentas
        for idx, (cuenta, sub_list) in enumerate(cuentas.items(), start=1):
            sheet_name = str(idx)

            if sheet_name not in workbook.sheetnames:
                print(f"⚠️ [5.4] La hoja '{sheet_name}' no existe en la plantilla; se omite.")
                continue

            sh = workbook[sheet_name]

            # 🔹 APLICAR FECHAS SOLO PARA ESTA CÉDULA (sin tocar lógica global)
            for clave, (col_letra, fila) in CELDAS_FECHAS_54.items():
                fecha = fechas.get(clave)
                if not fecha:
                    continue
                sh[f"{col_letra}{fila}"].value = fecha.strftime("%d/%m/%Y")

            # 🔹 Limpiar el área de detalle (solo filas de subcuentas, NO la de totales)
            for row in range(AREA_DETALLE_INICIO, AREA_DETALLE_FIN + 1):
                for col in range(1, 10):  # A..I
                    sh.cell(row=row, column=col).value = None

            # Título de la cédula: preservamos prefijo original si existe
            titulo_original = sh["A11"].value or ""
            prefijo = titulo_original.split(":")[0] if ":" in titulo_original else titulo_original
            if prefijo:
                sh["A11"].value = f"{prefijo}: CÉDULA SUMARIA - {cuenta}"
            else:
                sh["A11"].value = f"CÉDULA SUMARIA - {cuenta}"

            row = AREA_DETALLE_INICIO  # primera fila de detalle

            for i, sub in enumerate(sub_list, start=1):
                sh[f"A{row}"].value = i                         # Código correlativo
                sh[f"B{row}"].value = sub.nombre_subcuenta      # Nombre subcuenta
                sh[f"C{row}"].value = ""                        # Ref P/T vacío

                # Montos
                sh[f"D{row}"].value = sub.saldo_ini_anterior or 0
                sh[f"E{row}"].value = sub.saldo_julio_anterior or 0
                sh[f"F{row}"].value = sub.saldo_dic_anterior or 0
                sh[f"G{row}"].value = sub.ajuste_debe or 0
                sh[f"H{row}"].value = sub.ajuste_haber or 0
                sh[f"I{row}"].value = sub.saldo_dic_actual or 0

                row += 1


        print("✔️ [5.4] Cédulas 5.4 OTRAS CÉDULAS SUMARIAS (EstadoResultados) actualizadas ✔️")
        return workbook

    except Exception as e:
        print("⚠️ Error al actualizar cédulas 5.4 OTRAS CÉDULAS SUMARIAS (EstadoResultados):", e)
        return workbook
