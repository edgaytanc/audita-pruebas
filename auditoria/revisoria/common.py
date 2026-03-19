# auditoria/revisoria_common.py

import unicodedata
from openpyxl.cell.cell import MergedCell

from auditoria.models import BalanceCuentas


# ================== UTILIDADES TEXTO ==================

def _clean(value):
    return str(value).lower().replace(" ", "").replace("\xa0", "") if value else ""


def _normalizar(value):
    if not value:
        return ""
    value = ''.join(
        c for c in unicodedata.normalize('NFD', str(value))
        if unicodedata.category(c) != 'Mn'
    )
    return value.lower().replace(" ", "").replace("\xa0", "")


# ================== UTIL PARA CELDAS (MERGED) ==================

def _set_cell_value(sheet, row, col, value):
    """
    Escribe respetando celdas combinadas: si (row,col) está dentro de un merge,
    escribe en la celda superior izquierda del rango.
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in sheet.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                sheet.cell(rng.min_row, rng.min_col).value = value
                return
    cell.value = value


# ================== FECHAS ==================

def _formatear_fecha(fecha_obj):
    """Devuelve texto tipo '31/12/2024'."""
    if not fecha_obj:
        return None
    return fecha_obj.strftime("%d/%m/%Y")


def _extraer_fechas_desde_db(audit_id):
    """
    Fallback: toma fechas desde BalanceCuentas (Activo, ANUAL)
    y las clasifica en d1..d4.

    1) Mantiene la lógica original:
       - d1:  01/12/xxxx
       - d2:  31/07/xxxx
       - d3/d4: todas las 31/12 (primera = d3, última = d4)
    2) Si alguna de d1..d4 queda vacía (porque se usaron
       otras fechas distintas), rellena los huecos con las
       fechas que haya en BD, sin tocar lo que ya se encontró.
    3) Caso especial (SOLO para tu escenario): columnas E y H pueden tener
       la MISMA fecha. En importación se distinguen por tipo_cuenta:
         - E  -> NT_ANT
         - H  -> NT_ACT
       En ese caso d3/d4 se toman directamente de esas etiquetas (aunque sean iguales).
    """

    qs_fechas = BalanceCuentas.objects.filter(
        audit_id=audit_id,
        seccion__iexact="Activo",
        tipo_balance="ANUAL"
    ).values_list("fecha_corte", flat=True).distinct()

    fechas = sorted(set(qs_fechas))

    res = {"d1": None, "d2": None, "d3": None, "d4": None}
    if not fechas:
        return res

    # -----------------------------------------------------------
    # 0) CASO ESPECIAL ROBUSTO: d3/d4 desde NT_ANT y NT_ACT (E/H)
    #    (esto permite que d3 == d4 cuando E y H son iguales)
    # -----------------------------------------------------------
    tagged = BalanceCuentas.objects.filter(
        audit_id=audit_id,
        seccion__iexact="Activo",
        tipo_balance="ANUAL",
        tipo_cuenta__in=["NT_ANT", "NT_ACT"],
    ).values_list("tipo_cuenta", "fecha_corte").distinct()

    tagged_map = {}
    for tc, fc in tagged:
        if tc and fc:
            tagged_map[str(tc).strip().upper()] = fc

    if "NT_ANT" in tagged_map and "NT_ACT" in tagged_map:
        # d3 = E, d4 = H (aunque sean la misma fecha)
        res["d3"] = tagged_map["NT_ANT"]
        res["d4"] = tagged_map["NT_ACT"]

    for f in fechas:
        # d1: 01/12/AAAA
        if f.month == 12 and f.day == 1 and res["d1"] is None:
            res["d1"] = f
        # d2: 31/07/AAAA
        elif f.month == 7 and f.day == 31 and res["d2"] is None:
            res["d2"] = f

    if res["d3"] is None or res["d4"] is None:
        de_31_dic = [f for f in fechas if f.month == 12 and f.day == 31]
        de_31_dic.sort()
        if de_31_dic and res["d3"] is None:
            res["d3"] = de_31_dic[0]
        if de_31_dic and len(de_31_dic) > 1 and res["d4"] is None:
            res["d4"] = de_31_dic[-1]

    # ---------- 2) FALLBACK SOLO SI FALTAN FECHAS ----------
    usadas = {f for f in res.values() if f is not None}

    restantes = [f for f in fechas if f not in usadas]
    restantes.sort()

    if res["d4"] is None and restantes:
        res["d4"] = restantes[-1]
        usadas.add(res["d4"])
        restantes = [f for f in restantes if f != res["d4"]]

    if res["d3"] is None and restantes:
        if res["d4"]:
            menores = [f for f in restantes if f < res["d4"]]
            if menores:
                res["d3"] = menores[-1]
                usadas.add(res["d3"])
                restantes = [f for f in restantes if f != res["d3"]]
        if res["d3"] is None and restantes:
            res["d3"] = restantes[0]
            usadas.add(res["d3"])
            restantes = [f for f in restantes if f != res["d3"]]

    if res["d2"] is None and restantes:
        res["d2"] = restantes[0]
        usadas.add(res["d2"])
        restantes = [f for f in restantes if f != res["d2"]]

    if res["d1"] is None and restantes:
        res["d1"] = restantes[0]

    return res


def _aplicar_fechas_a_balance(sheet_balance, fechas):
    """
    Escribe las fechas en las posiciones fijas de Hoja1 CENTRALIZADORA:
      d1 → E12
      d2 → F12
      d3 → G12
      d4 → J12
    """
    posiciones = {
        "d1": (12, 5),   # E12
        "d2": (12, 6),   # F12
        "d3": (12, 7),   # G12
        "d4": (12, 10),  # J12
    }

    for key, (fila, col) in posiciones.items():
        f = fechas.get(key)
        if f:
            texto = _formatear_fecha(f)
            _set_cell_value(sheet_balance, fila, col, texto)
        else:
            print(f"⚠ Fecha {key} no encontrada → no se aplica en CENTRALIZADORA ({fila}, {col})")


def _aplicar_fechas_a_estados_financieros(sheet_estados, fechas):
    """
    Pone las fechas en la fila de encabezados de la hoja
    'ESTADOS FINANCIEROS …' (C6, D6, E6, H6).
    """
    posiciones = {
        "d1": (6, 3),   # C6
        "d2": (6, 4),   # D6
        "d3": (6, 5),   # E6
        "d4": (6, 8),   # H6
    }

    for key, (fila, col) in posiciones.items():
        f = fechas.get(key)
        if f:
            texto = _formatear_fecha(f)
            _set_cell_value(sheet_estados, fila, col, texto)
        else:
            print(f"⚠ Fecha {key} no encontrada para ESTADOS ({fila}, {col})")
