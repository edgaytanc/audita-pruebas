# auditoria/revisoria_activo.py  (o el nombre que ya tengas)

from datetime import date

from auditoria.models import BalanceCuentas
from .common import (
    _clean,
    _normalizar,
    _extraer_fechas_desde_db,
    _aplicar_fechas_a_balance,
    _aplicar_fechas_a_estados_financieros,
)

from .pasivo_patrimonio import (
    _obtener_cuentas_pasivo_patrimonio,
    _obtener_ajustes_pasivo_patrimonio,
    _actualizar_bloque_pasivo_patrimonio,
    _actualizar_hoja_estados_financieros_pasivo_patrimonio,
    _obtener_cuentas_pasivo,
    _obtener_ajustes_pasivo,
    _actualizar_bloque_pasivo_centralizadora,
    _encontrar_fila_codigo_pasivo_patrimonio,
    _encontrar_fila_codigo_pasivo,
)

from openpyxl import load_workbook
from auditoria.processors.revisoria_cedulas import (
    llenar_cedula_31_efectivo_y_equivalentes,
    llenar_cedula_33_cuentas_por_cobrar,
    llenar_cedula_34_inventarios,
)
from audits.models import Audit
from auditoria.models import BalanceCuentas, RevisoriaSubcuenta
from collections import defaultdict


# ================== BUSCAR FILAS CLAVE ==================

def _encontrar_fila_codigo_activo(sheet):
    """
    Busca la fila donde aparece el encabezado 'Código+Activo'
    en las columnas B o C.
    """
    for row in range(1, sheet.max_row + 1):
        for col in (2, 3):
            if any(x in _normalizar(sheet.cell(row=row, column=col).value)
                   for x in ["codigo+activo", "codigo+activos"]):
                return row
    return None


def _encontrar_fila_suma_activo(sheet, fila_inicio):
    """
    Busca la fila donde está 'Suma Activo' en las columnas B o C,
    a partir de fila_inicio.
    """
    for row in range(fila_inicio + 1, sheet.max_row + 1):
        for col in (2, 3):
            if "sumaactivo" in _clean(sheet.cell(row=row, column=col).value):
                return row
    return None


# ================== OBTENER DATOS DB (CUENTAS) ==================

def _obtener_cuentas_activo(audit_id):
    """
    Arma una estructura por cuenta con d1, d2, d3, d4
    usando BalanceCuentas (tipo_balance = ANUAL, seccion = Activo).

    Ahora usa exactamente las mismas fechas que calcula `_extraer_fechas_desde_db`
    para que los encabezados y los valores vayan alineados.
    """

    # 1) Sacamos las fechas d1..d4 desde BD
    fechas = _extraer_fechas_desde_db(audit_id)

    d1 = fechas.get("d1")
    d2 = fechas.get("d2")
    d3 = fechas.get("d3")
    d4 = fechas.get("d4")

    qs = BalanceCuentas.objects.filter(
        audit_id=audit_id,
        seccion__iexact="Activo",
        tipo_balance="ANUAL",
    ).order_by("nombre_cuenta", "fecha_corte")

    cuentas = {}

    for bc in qs:
        nombre = (bc.nombre_cuenta or "").strip()
        if not nombre:
            continue

        f = bc.fecha_corte
        slot = None

        # 🔗 Mapeo 1:1 contra las fechas clasificadas
        tc = (bc.tipo_cuenta or "").strip().upper()
        if d1 and f == d1:
            slot = "d1"
        elif d2 and f == d2:
            slot = "d2"
        # ✅ Si d3 y d4 son la MISMA fecha, distinguir por tipo_cuenta
        elif d3 and d4 and d3 == d4 and f == d3:
            slot = "d4" if tc == "NT_ACT" else "d3"
        elif d3 and f == d3:
            slot = "d3"
        elif d4 and f == d4:
            slot = "d4"


        if not slot:
            # Fecha que no coincide con ninguno de los cortes oficiales
            continue

        if nombre not in cuentas:
            cuentas[nombre] = {
                "nombre": nombre,
                "d1": None,
                "d2": None,
                "d3": None,
                "d4": None,
            }

        cuentas[nombre][slot] = float(bc.valor)

    # Resumen por columna, para verificar que d1 tiene datos
    totales_por_slot = {"d1": 0, "d2": 0, "d3": 0, "d4": 0}
    for c in cuentas.values():
        for slot in totales_por_slot.keys():
            if c[slot] is not None:
                totales_por_slot[slot] += 1

    return list(cuentas.values())


def _obtener_ajustes_activo(audit_id):
    """
    Devuelve un dict con Debe/Haber por cuenta para los AJUSTES de Activo.
    {
        'Casa': {'debe': 150000.0, 'haber': 120000.0},
        ...
    }
    """
    qs = BalanceCuentas.objects.filter(
        audit_id=audit_id,
        seccion__iexact="Activo",
        tipo_balance="AJUSTE"
    )

    ajustes = {}
    for bc in qs:
        nombre = bc.nombre_cuenta
        if nombre not in ajustes:
            ajustes[nombre] = {'debe': None, 'haber': None}

        if (bc.tipo_cuenta or "").upper() == "DEBE":
            ajustes[nombre]['debe'] = float(bc.valor)
        elif (bc.tipo_cuenta or "").upper() == "HABER":
            ajustes[nombre]['haber'] = float(bc.valor)

    return ajustes


# ================== ESCRIBIR EN EXCEL (ACTIVO CENTRALIZADORA) ==================

def _actualizar_bloque_activo(sheet, audit_id, cuentas=None, ajustes=None):
    """
    Actualiza el bloque ACTIVO en la hoja CENTRALIZADORA (Hoja1).
    También rellena Debe/Haber si existen ajustes.
    """

    fila_header = _encontrar_fila_codigo_activo(sheet)
    fila_suma = _encontrar_fila_suma_activo(sheet, fila_header)

    if cuentas is None:
        cuentas = _obtener_cuentas_activo(audit_id)
    if ajustes is None:
        ajustes = _obtener_ajustes_activo(audit_id)

    if not fila_header or not fila_suma or not cuentas:
        print("⚠ No se pudo localizar bloque ACTIVO (header/suma/cuentas faltan)")
        return

    fila_inicio = fila_header + 1
    filas_actuales = fila_suma - fila_inicio
    filas_necesarias = len(cuentas)

    # Ajustar filas dinámicamente
    if filas_necesarias > filas_actuales:
        sheet.insert_rows(fila_suma, filas_necesarias - filas_actuales)
    elif filas_necesarias < filas_actuales:
        for r in range(fila_inicio + filas_necesarias, fila_suma):
            for c in range(1, 20):
                sheet.cell(row=r, column=c).value = None

    # Columnas (según tu CENTRALIZADORA)
    COL_COD = 2     # Código+Activo
    COL_NOM = 3     # Cuenta
    COL_D1 = 5      # Al 01/12/AAAA  (nueva columna)
    COL_D2 = 6      # Al 31/07/AAAA
    COL_D3 = 7      # Al 31/12 año anterior
    COL_DEBE = 8    # Debe (ajustes)
    COL_HABER = 9   # Haber (ajustes)
    COL_D4 = 10     # Al 31/12 año actual

    for i, c in enumerate(cuentas):
        r = fila_inicio + i
        nombre = c["nombre"]

        # Valores que vamos a escribir (para log)
        v_d1 = c.get("d1")
        v_d2 = c.get("d2")
        v_d3 = c.get("d3")
        v_d4 = c.get("d4")

        # 🔧 PARCHE: si d1 viene vacío pero d4 tiene valor, usar d4 también en la col E
        if v_d1 is None and v_d4 is not None:
            v_d1 = v_d4

        datos_ajuste = ajustes.get(nombre, {}) if ajustes else {}
        v_debe = datos_ajuste.get("debe")
        v_haber = datos_ajuste.get("haber")


        # Escribir base
        sheet.cell(row=r, column=COL_COD).value = i + 1
        sheet.cell(row=r, column=COL_NOM).value = nombre
        sheet.cell(row=r, column=COL_D1).value = v_d1
        sheet.cell(row=r, column=COL_D2).value = v_d2
        sheet.cell(row=r, column=COL_D3).value = v_d3
        sheet.cell(row=r, column=COL_D4).value = v_d4

        # Debe / Haber desde ajustes (si existen)
        if datos_ajuste:
            sheet.cell(row=r, column=COL_DEBE).value = v_debe
            sheet.cell(row=r, column=COL_HABER).value = v_haber

    print("✔️ ACTIVO → Base actualizada (CENTRALIZADORA)")


# ================== ACTUALIZAR HOJA "ESTADOS FINANCIEROS" ==================

def _actualizar_hoja_estados_financieros(workbook, cuentas, ajustes, fechas):
    """
    Replica los mismos datos de ACTIVO en la hoja 'ESTADOS FINANCIEROS…'
    si existe, usando siempre las mismas filas/columnas de tu plantilla.
    También actualiza las fechas del encabezado y Debe/Haber.
    """
    # Buscar hoja cuyo nombre contenga "ESTADOS FINANCIEROS"
    target_name = None
    for name in workbook.sheetnames:
        if "ESTADOS FINANCIEROS" in str(name).upper():
            target_name = name
            break

    if not target_name:
        print("ℹ Hoja 'ESTADOS FINANCIEROS …' no encontrada, se omite.")
        return

    sheet = workbook[target_name]
    print(f"🔄 Actualizando hoja {target_name}…")

    # Fechas en encabezado
    _aplicar_fechas_a_estados_financieros(sheet, fechas)

    FILA_INICIO = 7

    # Columnas:
    COL_COD = 1   # A → Código+Activo
    COL_NOM = 2   # B → Cuenta
    COL_D1 = 3    # C → Fecha corte 1
    COL_D2 = 4    # D → Fecha corte 2
    COL_D3 = 5    # E → Fecha corte 3
    COL_DEBE = 6  # F → Debe
    COL_HABER = 7 # G → Haber
    COL_D4 = 8    # H → Fecha corte 4 (año actual)

    for i, c in enumerate(cuentas):
        r = FILA_INICIO + i
        nombre = c["nombre"]

        sheet.cell(row=r, column=COL_COD).value = i + 1
        sheet.cell(row=r, column=COL_NOM).value = nombre
        sheet.cell(row=r, column=COL_D1).value = c["d1"]
        sheet.cell(row=r, column=COL_D2).value = c["d2"]
        sheet.cell(row=r, column=COL_D3).value = c["d3"]
        sheet.cell(row=r, column=COL_D4).value = c["d4"]

        datos_ajuste = ajustes.get(nombre, {})
        if datos_ajuste:
            sheet.cell(row=r, column=COL_DEBE).value = datos_ajuste.get("debe")
            sheet.cell(row=r, column=COL_HABER).value = datos_ajuste.get("haber")

    print("✔️ ACTIVO → Base actualizada (ESTADOS FINANCIEROS)")


# ================== FLUJO PRINCIPAL ==================

def update_dynamic_excel(workbook, sheet_name, audit_id, sheet_financiero=None):
    """
    MODO 1 → 2 CENTRALIZADORA BALANCE.xlsx
        1) Actualiza bloque ACTIVO (códigos, cuentas, d1..d4) en CENTRALIZADORA (Hoja1).
        2) Actualiza encabezados de fechas en E12, F12, G12, J12 usando BD.
        3) Replica esos mismos datos + Debe/Haber en la hoja de ESTADOS FINANCIEROS
           (Revisoria) si existe.
        4) Actualiza bloque CÓDIGO+PASIVO+PATRIMONIO.

    MODO 2 → 4 CENTRALIZADORA DEL PASIVO.xlsx
        - Solo actualiza bloque CÓDIGO+PASIVO (sin tocar ACTIVO ni PASIVO+PATRIMONIO).
    """
    sheet_balance = workbook[sheet_name]

    # ============================================================
    # 0️⃣ Detectar qué tipo de plantilla es según los encabezados
    # ============================================================
    # Estas funciones ya existen en tu módulo:
    #   _encontrar_fila_codigo_activo
    #   _encontrar_fila_codigo_pasivo_patrimonio
    # Las nuevas que te pasé:
    #   _encontrar_fila_codigo_pasivo

    fila_activo = _encontrar_fila_codigo_activo(sheet_balance)
    fila_pasivo_patr = _encontrar_fila_codigo_pasivo_patrimonio(sheet_balance)
    fila_pasivo = _encontrar_fila_codigo_pasivo(sheet_balance)

    # Caso B → Plantilla SOLO PASIVO (4 CENTRALIZADORA DEL PASIVO.xlsx)
    # No tiene "Código+Activo" ni "Código+Pasivo+Patrimonio", pero sí "Código+Pasivo".
    if fila_activo is None and fila_pasivo_patr is None and fila_pasivo is not None:

        cuentas_pasivo = _obtener_cuentas_pasivo(audit_id)
        ajustes_pasivo = _obtener_ajustes_pasivo(audit_id)

        if not cuentas_pasivo:
            print("⚠ No hay cuentas de PASIVO en BD para esta auditoría.")
            return workbook

        _actualizar_bloque_pasivo_centralizadora(
            sheet_balance,
            audit_id,
            cuentas=cuentas_pasivo,
            ajustes=ajustes_pasivo,
        )

        # 🔸 IMPORTANTE: en esta centralizadora las fechas ya las maneja la plantilla
        # (fórmulas), por eso NO escribimos encabezados de fecha aquí.
        print("✔️ CENTRALIZADORA DEL PASIVO actualizada ✔️")
        return workbook

    # ============================================================
    # Caso A → Plantilla completa de BALANCE (2 CENTRALIZADORA BALANCE.xlsx)
    # ============================================================
    # 1️⃣ Traer cuentas + ajustes SOLO de ACTIVO
    cuentas, ajustes = _obtener_cuentas_y_ajustes_activo(audit_id)

    # Si no hay cuentas, no hacemos nada
    if not cuentas:
        print("⚠ No hay cuentas de ACTIVO en BD para esta auditoría.")
        return workbook

    # 2️⃣ Bloque ACTIVO en CENTRALIZADORA
    try:
        _actualizar_bloque_activo(sheet_balance, audit_id, cuentas=cuentas, ajustes=ajustes)
    except Exception as e:
        # Si por alguna razón no hay bloque ACTIVO en esta hoja, no queremos romper
        print(f"ℹ No se pudo actualizar bloque ACTIVO en esta hoja: {e}")

    # 3️⃣ Fechas desde BD (ANUAL, ACTIVO)
    fechas = _extraer_fechas_desde_db(audit_id)

    try:
        _aplicar_fechas_a_balance(sheet_balance, fechas)
    except Exception as e:
        print(f"ℹ No se pudieron aplicar fechas en encabezado de ACTIVO: {e}")

    # 4️⃣ Replicar datos en hoja ESTADOS FINANCIEROS (Revisoria) – ACTIVO
    try:
        _actualizar_hoja_estados_financieros(workbook, cuentas, ajustes, fechas)
    except Exception as e:
        print(f"⚠ Error actualizando hoja ESTADOS FINANCIEROS: {e}")

    # ========= PASIVO + PATRIMONIO =========
    cuentas_pp = _obtener_cuentas_pasivo_patrimonio(audit_id)
    ajustes_pp = _obtener_ajustes_pasivo_patrimonio(audit_id)

    if cuentas_pp:
        _actualizar_bloque_pasivo_patrimonio(
            sheet_balance,
            audit_id,
            cuentas=cuentas_pp,
            ajustes=ajustes_pp,
        )

        try:
            _actualizar_hoja_estados_financieros_pasivo_patrimonio(
                workbook,
                cuentas_pp,
                ajustes_pp,
            )
        except Exception as e:
            print(f"⚠ Error actualizando hoja ESTADOS FINANCIEROS (PASIVO+PATRIMONIO): {e}")
    else:
        print("ℹ No hay cuentas de PASIVO/PATRIMONIO para esta auditoría.")

    print("✔️ Balance dinámico COMPLETADO ✔️")
    return workbook



def _obtener_cuentas_y_ajustes_activo(audit_id):
    """
    Devuelve:
      - cuentas: lista de dicts {nombre, d1, d2, d3, d4}
      - ajustes: dict { nombre_cuenta: {"debe": float, "haber": float} }
    Solo se toma seccion = 'Activo'.

    Ahora reutiliza el mismo mapeo de fechas que _obtener_cuentas_activo
    para que las columnas E/F/G/J de CENTRALIZADORA y la hoja ESTADOS
    queden alineadas con las fechas d1..d4.
    """
    cuentas = _obtener_cuentas_activo(audit_id)
    ajustes = _obtener_ajustes_activo(audit_id)


    return cuentas, ajustes

# ================== OBTENER DATOS SOLO DE PASIVO ==================

def update_cedula_31_efectivo(workbook, sheet_name, audit_id, sheet_financiero=None):
    """
    Actualiza el archivo '3.1 EFECTIVO_Y_EQUIVALENTES.xlsx'.

    - Usa las subcuentas de RevisoriaSubcuenta para la cuenta
      'Efectivo y Equivalentes'.
    - Escribe en la hoja indicada por `sheet_name` (debería ser 'A-SUM').
    """
    sheet = workbook[sheet_name]   # normalmente 'A-SUM'
    audit = Audit.objects.get(pk=audit_id)

    # 👉 Aquí usamos la función que ya creaste en revisoria_cedulas.py
    llenar_cedula_31_efectivo_y_equivalentes(sheet, audit)

    print("✔️ Cédula 3.1 EFECTIVO Y EQUIVALENTES actualizada ✔️")
    return workbook

def update_cedula_33_cuentas_por_cobrar(workbook, sheet_name, audit_id, sheet_financiero=None):
    """
    Actualiza el archivo '3.3 CUENTAS_POR_COBRAR.xlsx'.

    - Usa las subcuentas de RevisoriaSubcuenta para la cuenta
      'Cuentas por Cobrar'.
    - Escribe en la hoja indicada por `sheet_name`
      (debería ser 'C CUENTAS POR COBRAR').
    """
    sheet = workbook[sheet_name]   # normalmente 'C CUENTAS POR COBRAR'
    audit = Audit.objects.get(pk=audit_id)

    llenar_cedula_33_cuentas_por_cobrar(sheet, audit)

    print("✔️ Cédula 3.3 CUENTAS POR COBRAR actualizada ✔️")
    return workbook

def update_cedula_34_inventarios(workbook, sheet_name, audit_id, sheet_financiero=None):
    """
    Actualiza el archivo '3.4 INVENTARIOS.xlsx'.

    - Usa las subcuentas de RevisoriaSubcuenta para la cuenta
      'Inventarios'.
    - Escribe en la hoja indicada por `sheet_name` (debería ser 'D-SUM').
    """
    sheet = workbook[sheet_name]   # normalmente 'D-SUM'
    audit = Audit.objects.get(pk=audit_id)

    # 👉 Usamos la función de processors
    llenar_cedula_34_inventarios(sheet, audit)

    print("✔️ Cédula 3.4 INVENTARIOS actualizada ✔️")
    return workbook


def update_cedula_36_otras(workbook, audit_id, sheet_financiero=None):
    """
    Rellena la plantilla 3.6 OTRAS CÉDULAS SUMARIAS para ACTIVO.

    - Solo secciones que contengan 'ACTIVO'
    - Excluye las cuentas especiales (Caja, CxC, Efectivo, Inventarios, Inversiones, PPE)
      que ya tienen sus propias cédulas.
    - Cada cuenta_principal "otra" va a una hoja: '1', '2', '3', ...
    """

    from auditoria.models import RevisoriaSubcuenta

    # ⬅️ AQUÍ CAMBIAS LAS POSICIONES DE LAS FECHAS CUANDO QUIERAS
    # Por defecto: d1, d2, d3 en la fila 13 (encabezado) y d4 a la derecha.
    CELDAS_FECHAS_36 = {
        "d1": ("D", 13),   # 01/12/2025
        "d2": ("E", 13),   # 31/07/2023
        "d3": ("F", 13),   # 31/12/2022
        "d4": ("I", 13),   # 31/12/2024 (o donde la tengas)
    }

    # 🔹 rango de filas SOLO para subcuentas (la 29 se deja para las sumas)
    AREA_DETALLE_INICIO = 14  # primera fila de detalle
    AREA_DETALLE_FIN    = 28  # última fila de detalle (29 = totales, no se toca)

    CUENTAS_ESPECIALES_ACTIVO = {
        "Efectivo y Equivalentes",
        "Inversiones",
        "Cuentas por Cobrar",
        "Inventarios",
        "Propiedad, Planta y Equipo",
    }

    try:
        # Todas las subcuentas de ACTIVO que sí van a 3.6 (otras cuentas)
        qs = (
            RevisoriaSubcuenta.objects
            .filter(audit_id=audit_id, seccion__icontains="ACTIVO")
            .exclude(cuenta_principal__in=CUENTAS_ESPECIALES_ACTIVO)
            .order_by("cuenta_principal", "id")
        )

        if not qs.exists():
            print("ℹ️ [3.6] No hay subcuentas 'otras' de ACTIVO; se omite 3.6.")
            return workbook

        # 🔹 Fechas (misma lógica que CENTRALIZADORA, pero posiciones propias)
        fechas = _extraer_fechas_desde_db(audit_id)

        def es_placeholder_x(nombre: str) -> bool:
            """
            Devuelve True si el nombre es algo tipo 'XXXX', 'XXXXXXXXX', etc.
            (solo X y espacios/guiones).
            """
            if not nombre:
                return False
            limpio = nombre.replace(" ", "").replace("-", "").upper()
            return len(limpio) > 0 and set(limpio) == {"X"}

        # Agrupar solo las filas que realmente son subcuentas de detalle
        cuentas = {}
        for sub in qs:
            cp = (sub.cuenta_principal or "").strip()
            nombre = (sub.nombre_subcuenta or "").strip()

            if not nombre:
                continue

            # Saltar cabeceras (nombre igual a la cuenta principal)
            if nombre.upper() == cp.upper():
                continue

            # Saltar totales / placeholders tipo XXXX / XXXXXXXX
            if es_placeholder_x(nombre):
                continue

            cuentas.setdefault(cp, []).append(sub)

        # Rellenar las hojas 1, 2, 3,... con esas subcuentas
        for idx, (cuenta, sub_list) in enumerate(cuentas.items(), start=1):
            sheet_name = str(idx)

            if sheet_name not in workbook.sheetnames:
                print(f"⚠️ [3.6] La hoja '{sheet_name}' no existe en la plantilla; se omite.")
                continue

            sh = workbook[sheet_name]

            # 🔹 APLICAR FECHAS SOLO PARA ESTA CÉDULA (sin tocar la lógica global)
            for clave, (col_letra, fila) in CELDAS_FECHAS_36.items():
                fecha = fechas.get(clave)
                if not fecha:
                    continue
                # Las demás funciones usan formato dd/mm/yyyy, replicamos eso:
                sh[f"{col_letra}{fila}"].value = fecha.strftime("%d/%m/%Y")

            # 🔹 Limpiar el área de detalle (solo filas de subcuentas, NO la de totales)
            for row in range(AREA_DETALLE_INICIO, AREA_DETALLE_FIN + 1):
                for col in range(1, 10):  # A..I
                    sh.cell(row=row, column=col).value = None

            # Título de la cédula
            sh["A11"].value = f"A-SUM: CÉDULA SUMARIA - {cuenta}"

            row = AREA_DETALLE_INICIO  # primera fila de detalle

            for i, sub in enumerate(sub_list, start=1):
                sh[f"A{row}"].value = i                         # Código correlativo
                sh[f"B{row}"].value = sub.nombre_subcuenta      # Nombre subcuenta
                sh[f"C{row}"].value = ""                        # Ref P/T vacío

                # Montos
                sh[f"D{row}"].value = sub.saldo_ini_anterior or 0
                sh[f"E{row}"].value = sub.saldo_julio_anterior or 0
                sh[f"F{row}"].value = sub.saldo_dic_anterior or 0
                sh[f"G{row}"].value = sub.ajuste_debe or 0
                sh[f"H{row}"].value = sub.ajuste_haber or 0
                sh[f"I{row}"].value = sub.saldo_dic_actual or 0

                row += 1

        return workbook

    except Exception as e:
        print("⚠️ Error al actualizar cédulas 3.6 OTRAS CÉDULAS SUMARIAS:", e)
        return workbook
