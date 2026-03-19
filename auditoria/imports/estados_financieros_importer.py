# auditoria/imports/estados_financieros_importer.py  (por ejemplo)

import openpyxl
from datetime import datetime
from django.db import transaction
from audits.models import Audit
from auditoria.models import BalanceCuentas, RegistroAuxiliar, SaldoInicial
import traceback
import io

# Importadores delegados
from .processors.annual_importer import process_annual_sheet as _process_annual_sheet
from .processors.semestral_importer import process_semestral_sheet as _process_semestral_sheet
from .processors.auxiliary_importer import process_auxiliary_records as _process_auxiliary_records
from .processors.initial_balances_importer import process_initial_balances as _process_initial_balances
from .processors.revisoria_importer import (
    process_revisoria_sheet as _process_revisoria_sheet,
    process_estado_resultados_revisoria as _process_estado_resultados_revisoria,
)
from auditoria.models import RevisoriaSubcuenta


class EstadosFinancierosImporter:
    """
    Importador unificado para archivos de estados financieros que contiene
    tanto balances anuales como semestrales en diferentes hojas.
    
    El archivo debe tener:
    - Una hoja "ESTADOS FINANCIEROS ANUAL" con balances anuales (2 fechas)
    - Una hoja "ESTADOS FINANCIEROS SEMESTRALES" con balances semestrales (4 fechas)
    - Opcionalmente, hojas para registros auxiliares y saldos iniciales
    """
    def __init__(self, file_obj, audit_id):
        """
        Inicializa el importador con un archivo en memoria y un ID de auditoría
        
        Args:
            file_obj: Un objeto InMemoryUploadedFile o similar que puede ser leído directamente
            audit_id: ID de la auditoría asociada
        """
        self.file_obj = file_obj
        self.audit_id = audit_id

    def validate_file(self):
        """Valida que el archivo contenga al menos una de las hojas requeridas (ANUAL o SEMESTRAL)"""
        try:
            file_content = io.BytesIO(self.file_obj.read())
            self.file_obj.seek(0)
            
            wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
            sheet_names_lower = [name.lower() for name in wb.sheetnames]

            # ✅ SOLO validamos anual / semestral (como antes)
            if not any(
                ('anual' in name and 'semestral' not in name) or
                ('semestral' in name)
                for name in sheet_names_lower
            ):
                return False

            return True
        except Exception:
            return False

    @transaction.atomic
    def process_file(self):
        """Procesa todas las hojas relevantes del archivo (ANUAL/SEMESTRAL/AUX/SALDOS)"""
        try:
            file_content = io.BytesIO(self.file_obj.read())
            self.file_obj.seek(0)
            
            wb = openpyxl.load_workbook(file_content, data_only=True)
            sheet_names_lower = {name.lower(): name for name in wb.sheetnames}
            
            anual_processed = False
            semestral_processed = False
            
            # ▶ Hoja ANUAL
            anual_sheet_name = next(
                (sheet_names_lower[name] for name in sheet_names_lower
                 if "anual" in name and "semestral" not in name),
                None,
            )
            if anual_sheet_name:
                self.process_annual_sheet(wb[anual_sheet_name])
                anual_processed = True

            # ▶ Hoja SEMESTRAL
            semestral_sheet_name = next(
                (sheet_names_lower[name] for name in sheet_names_lower 
                 if "semestral" in name),
                None
            )
            if semestral_sheet_name:
                self.process_semestral_sheet(wb[semestral_sheet_name])
                semestral_processed = True

            # ▶ Registros auxiliares (si hay hoja auxiliar)
            aux_sheet_name = next(
                (sheet_names_lower[name] for name in sheet_names_lower 
                 if "auxiliar" in name),
                None
            )
            if aux_sheet_name:
                self.process_auxiliary_records(wb[aux_sheet_name])
            
            # ▶ Saldos iniciales (si hay hoja de saldos)
            initial_sheet_name = next(
                (sheet_names_lower[name] for name in sheet_names_lower 
                 if "saldo" in name),
                None
            )
            if initial_sheet_name:
                self.process_initial_balances(wb[initial_sheet_name])
            
            # Debe haber al menos ANUAL o SEMESTRAL
            if not (anual_processed or semestral_processed):
                return False, "❌ No se encontró ninguna hoja válida de estados financieros (ANUAL/SEMESTRAL)"
            
            return True, "✅ Importación completada exitosamente"
        except Exception as e:
            error_msg = f"Error procesando archivo: {str(e)}"
            traceback.print_exc()
            return False, error_msg

    # ------------------------------------------------------------------
    #  Wrappers que delegan a los importadores especializados
    # ------------------------------------------------------------------

    def process_annual_sheet(self, sheet):
        """Enrutador hacia `imports.annual_importer.process_annual_sheet`."""
        return _process_annual_sheet(sheet, self.audit_id)

    def process_semestral_sheet(self, sheet):
        """Enrutador hacia `imports.semestral_importer.process_semestral_sheet`."""
        return _process_semestral_sheet(sheet, self.audit_id)

    def process_auxiliary_records(self, sheet):
        """Enrutador hacia `imports.auxiliary_importer.process_auxiliary_records`."""
        return _process_auxiliary_records(sheet, self.audit_id)

    def process_initial_balances(self, sheet):
        """Enrutador hacia `imports.initial_balances_importer.process_initial_balances`."""
        return _process_initial_balances(sheet, self.audit_id)


def importar_estados_financieros_revisoria(file_obj, audit_id):
    """
    Flujo ESPECÍFICO para Revisoría:
      1) Opcionalmente usa EstadosFinancierosImporter para balances ANUAL/SEMESTRAL
         (si el archivo también trae esas hojas).
      2) Siempre intenta procesar la hoja 'ESTADOS FINANCIEROS' para llenar
         los datos propios de Revisoria (fecha, Debe, Haber, etc.).
    """

    # 1️⃣ Intentar correr el importador general (no rompe si el archivo no tiene ANUAL/SEMESTRAL)
    importer = EstadosFinancierosImporter(file_obj, audit_id)
    ok_main, msg_main = importer.process_file()

    # 2️⃣ Releer archivo para procesar la hoja de Revisoria
    file_obj.seek(0)
    file_content = io.BytesIO(file_obj.read())
    file_obj.seek(0)

    wb = openpyxl.load_workbook(file_content, data_only=True)
    sheet_names_lower = {name.lower(): name for name in wb.sheetnames}

    revisoria_sheet_name = next(
        (sheet_names_lower[name] for name in sheet_names_lower
         if "estados financieros" in name),  # "ESTADOS FINANCIEROS"
        None
    )

    if not revisoria_sheet_name:
        # Si no hay hoja de Revisoria, devolvemos el resultado del importador general
        if ok_main:
            return True, msg_main
        return False, "❌ No se encontró hoja 'ESTADOS FINANCIEROS' para Revisoria"

    # 3️⃣ Procesar la hoja de Revisoria
    _process_revisoria_sheet(wb[revisoria_sheet_name], audit_id)
        # 3️⃣.bis Procesar también el bloque de ESTADO DE RESULTADOS
    try:
        _process_estado_resultados_revisoria(wb[revisoria_sheet_name], audit_id)
    except Exception as e:
        print(f"⚠️ Error importando bloque ESTADO DE RESULTADOS: {e}")

    
    # 3.1️⃣ 👇 NUEVO: importar subcuentas desde las subhojas
    importar_subcuentas_estados_financieros(audit_id, wb)

    # 4️⃣ Mensajes combinados
    if ok_main:
        return True, "✅ Importación financiera/interna completada + Revisoria cargada correctamente"
    else:
        # La parte ANUAL/SEMESTRAL pudo fallar, pero Revisoria sí se cargó
        return True, f"⚠ Revisoria cargada. Detalle importación general: {msg_main}"

class RevisoriaEstadosFinancierosImporter:
    """
    Importador exclusivo para auditorías de Revisoría Fiscal.

    - NO toca la lógica de financiera/interna.
    - Solo busca una hoja tipo "ESTADOS FINANCIEROS" (o similar)
      y la procesa con `processors.revisoria_importer.process_revisoria_sheet`.
    """

    def __init__(self, file_obj, audit_id):
        """
        Igual que el importador general:
        - file_obj: archivo subido (InMemoryUploadedFile, etc.)
        - audit_id: id de la auditoría
        """
        self.file_obj = file_obj
        self.audit_id = audit_id

    def validate_file(self):
        """
        Valida que el archivo tenga una hoja que parezca de estados financieros
        para Revisoría (nombre con 'estados financieros' o 'revisoria').
        NO afecta al validador del importador general.
        """
        try:
            file_content = io.BytesIO(self.file_obj.read())
            self.file_obj.seek(0)

            wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
            sheet_names_lower = [name.lower() for name in wb.sheetnames]

            tiene_revisoria = any(
                ("estados financieros" in name) or
                ("revisoria" in name)
                for name in sheet_names_lower
            )
            return tiene_revisoria
        except Exception:
            return False

    @transaction.atomic
    def process_file(self):
        """
        Procesa SOLO la hoja de Revisoría Fiscal.
        Usa el procesador especializado `_process_revisoria_sheet`.
        """
        file_content = io.BytesIO(self.file_obj.read())
        self.file_obj.seek(0)

        wb = openpyxl.load_workbook(file_content, data_only=True)
        sheet_names_lower = {name.lower(): name for name in wb.sheetnames}

        # Buscar hoja de Revisoría: por nombre con 'estados financieros' o 'revisoria'
        revisoria_sheet_name = next(
            (
                sheet_names_lower[name]
                for name in sheet_names_lower
                if ("estados financieros" in name) or ("revisoria" in name)
            ),
            None,
        )

        if not revisoria_sheet_name:
            return False, "❌ No se encontró una hoja de ESTADOS FINANCIEROS para Revisoría Fiscal."

        # Procesar con el importador especializado
        sheet = wb[revisoria_sheet_name]
        _process_revisoria_sheet(sheet, self.audit_id)
        
        # 👇 NUEVO: también aquí importamos subcuentas
        importar_subcuentas_estados_financieros(self.audit_id, wb)

        return True, "✅ Importación de estados financieros de Revisoría completada exitosamente."

def _num(value):
    try:
        return float(value) if value not in (None, "") else None
    except Exception:
        return None


def importar_subcuentas_estados_financieros(audit_id, workbook):
    """
    Lee las hojas de SUBCUENTAS (ACTIVO / PASIVO / PATRIMONIO / RESULTADOS) y
    guarda subcuentas en RevisoriaSubcuenta.

    🔒 No toca BalanceCuentas ni nada de lo que ya funciona.
    """
    from audits.models import Audit
    audit = Audit.objects.get(pk=audit_id)

    # 1) Obtener las cuentas principales desde la hoja ESTADOS FINANCIEROS
    if "ESTADOS FINANCIEROS" not in workbook.sheetnames:
        # Sin hoja principal, no tiene sentido seguir
        return

    main_sheet = workbook["ESTADOS FINANCIEROS"]
    cuentas_principales = set()

    for row in main_sheet.iter_rows(min_row=3, values_only=True):
        codigo = row[0]
        cuenta = row[1]
        if not cuenta:
            continue
        # No quemamos códigos: solo tomamos cualquier fila que tenga cuenta
        cuentas_principales.add(str(cuenta).strip())

    # 2) Mapear nombres de hojas ignorando mayúsculas/espacios
    sheetnames_lower = {name.lower(): name for name in workbook.sheetnames}

    def _buscar_hoja(pattern_subcuentas, pattern_tipo):
        return next(
            (
                sheetnames_lower[name]
                for name in sheetnames_lower
                if pattern_subcuentas in name and pattern_tipo in name
            ),
            None,
        )

    hoja_sub_activo      = _buscar_hoja("subcuentas", "activo")
    hoja_sub_pasivo      = _buscar_hoja("subcuentas", "pasivo")
    hoja_sub_patrimonio  = _buscar_hoja("subcuentas", "patrimonio")
    hoja_sub_result      = _buscar_hoja("subcuentas", "resultados")

    if hoja_sub_activo:
        _procesar_subhoja_subcuentas(
            audit,
            workbook[hoja_sub_activo],
            seccion_fija="Activo",
            cuentas_principales=cuentas_principales,
        )

    if hoja_sub_pasivo:
        # En esta hoja se va moviendo entre PASIVO y PATRIMONIO (si llegara a usarse así)
        _procesar_subhoja_subcuentas(
            audit,
            workbook[hoja_sub_pasivo],
            seccion_fija=None,   # se detecta dentro
            cuentas_principales=cuentas_principales,
        )

    if hoja_sub_patrimonio:
        # Hoja exclusiva de Patrimonio (caso actual de tu Excel)
        _procesar_subhoja_subcuentas(
            audit,
            workbook[hoja_sub_patrimonio],
            seccion_fija="Patrimonio",
            cuentas_principales=cuentas_principales,
        )

    if hoja_sub_result:
        _procesar_subhoja_subcuentas(
            audit,
            workbook[hoja_sub_result],
            seccion_fija="EstadoResultado",
            cuentas_principales=cuentas_principales,
        )


def _procesar_subhoja_subcuentas(audit, sheet, seccion_fija, cuentas_principales):
    """
    Procesa una hoja de SUBCUENTAS genérica:
    - Si seccion_fija es 'Activo', 'Patrimonio' o 'EstadoResultado', usamos esa sección fija.
    - Si seccion_fija es None (caso PASIVO), detectamos PASIVO / PATRIMONIO
      según las filas que traen esos textos.
    - No se queman nombres de cuentas: se basa en 'cuentas_principales' y textos tipo 'TOTAL...'.
    """
    current_seccion = seccion_fija  # 'Activo', 'Patrimonio', 'EstadoResultado' o None
    current_cuenta_principal = None

    for row in sheet.iter_rows(min_row=5, values_only=True):
        nombre = row[0]
        if not nombre:
            continue

        nombre_str = str(nombre).strip()
        upper = nombre_str.upper()

        # Encabezados generales de bloque (ACTIVO, PASIVO, PATRIMONIO, ESTADO DE RESULTADOS)
        if upper in ("ACTIVO", "PASIVO", "PATRIMONIO", "ESTADO DE RESULTADOS"):
            if seccion_fija is None:
                # Solo si NO hay seccion fija cambiamos entre PASIVO / PATRIMONIO
                if upper == "PASIVO":
                    current_seccion = "Pasivo"
                elif upper == "PATRIMONIO":
                    current_seccion = "Patrimonio"
            current_cuenta_principal = None
            continue

        # Filas tipo 'TOTAL ...' cierran el bloque de cuenta principal
        if upper.startswith("TOTAL "):
            current_cuenta_principal = None
            continue

        # ¿Es una cuenta principal? (coincide con la hoja ESTADOS FINANCIEROS)
        if nombre_str in cuentas_principales:
            current_cuenta_principal = nombre_str
            continue

        # Si no tenemos sección o cuenta principal, no podemos ubicar esta fila
        if not current_seccion or not current_cuenta_principal:
            continue

        # A partir de aquí es SUBCUENTA real
        saldo_ini_ant    = _num(row[1])
        saldo_jul_ant    = _num(row[2])
        saldo_dic_ant    = _num(row[3])
        ajuste_debe      = _num(row[4])
        ajuste_haber     = _num(row[5])
        saldo_dic_actual = _num(row[6])

        RevisoriaSubcuenta.objects.update_or_create(
            audit=audit,
            seccion=current_seccion,
            cuenta_principal=current_cuenta_principal,
            nombre_subcuenta=nombre_str,
            defaults={
                "saldo_ini_anterior":   saldo_ini_ant,
                "saldo_julio_anterior": saldo_jul_ant,
                "saldo_dic_anterior":   saldo_dic_ant,
                "ajuste_debe":          ajuste_debe,
                "ajuste_haber":         ajuste_haber,
                "saldo_dic_actual":     saldo_dic_actual,
            },
        )