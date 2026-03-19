"""
Generador de Plantillas Excel Dinámico - Crea plantillas Excel personalizadas

⚠️ PROTOCOLO DE SEGURIDAD CRÍTICO:
- SOLO crea NUEVOS archivos Excel, NUNCA modifica plantillas existentes
- Todas las plantillas generadas son archivos temporales en memoria
- El usuario descarga un archivo NUEVO, las plantillas originales nunca se tocan
- Valida configuración antes de generar

ARQUITECTURA:
1. Recibe configuración (workbook_name, columns, symbols, options)
2. Crea nuevo workbook con openpyxl
3. Genera 3 hojas: Instructions, Template, Symbol Reference
4. Aplica formato profesional
5. Retorna archivo en memoria (BytesIO) para descarga
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class DynamicTemplateGenerator:
    """
    Generador de plantillas Excel dinámicas basadas en configuración de usuario.

    ⚠️ SEGURIDAD: Solo crea NUEVOS archivos, nunca modifica existentes.
    """

    # Colores profesionales para formato
    COLOR_HEADER_BG = 'D3D3D3'  # Gris claro
    COLOR_HEADER_TEXT = '0070C0'  # Azul profesional
    COLOR_INSTRUCTIONS_BG = 'FFF2CC'  # Amarillo claro
    COLOR_SYMBOL_BG = 'E7E6E6'  # Gris muy claro
    COLOR_DATA_BORDER = '000000'  # Negro

    def __init__(self, configuration):
        """
        Inicializar generador con configuración.

        Args:
            configuration: dict con estructura:
                {
                    'workbook_name': str,
                    'columns': [
                        {
                            'name': str,
                            'width': int,
                            'data_type': str,
                            'is_required': bool
                        }
                    ],
                    'symbols': [
                        {
                            'symbol': str,
                            'description': str,
                            'category': str
                        }
                    ],
                    'options': {
                        'include_headers': bool,
                        'include_instructions': bool,
                        'color_code_symbols': bool,
                        'add_data_validation': bool
                    },
                    'audit_info': {
                        'title': str,
                        'created_by': str
                    }
                }
        """
        self.config = configuration
        self.workbook = None
        self.validate_configuration()

    def validate_configuration(self):
        """
        Valida que la configuración sea correcta antes de generar.

        Raises:
            ValueError: Si la configuración es inválida
        """
        required_fields = ['workbook_name', 'columns', 'symbols']

        for field in required_fields:
            if field not in self.config:
                raise ValueError(f"Campo requerido faltante: {field}")

        if not self.config['columns'] or len(self.config['columns']) < 2:
            raise ValueError("Se requieren al menos 2 columnas")

        if not self.config['symbols'] or len(self.config['symbols']) < 30:
            raise ValueError("Se requieren al menos 30 símbolos")

        logger.debug(
            f"Configuración validada: "
            f"{len(self.config['columns'])} columnas, "
            f"{len(self.config['symbols'])} símbolos"
        )

    def generate(self):
        """
        Genera el archivo Excel completo con todas las hojas.

        ⚠️ SEGURIDAD: Crea un NUEVO workbook en memoria

        Returns:
            BytesIO: Archivo Excel en memoria listo para descarga

        Raises:
            Exception: Si ocurre algún error durante la generación
        """
        try:
            logger.info(
                f"Iniciando generación de plantilla: {self.config['workbook_name']}"
            )

            # ══════════════════════════════════════════════════════
            # PASO 1: Crear NUEVO workbook (nunca abre existente)
            # ══════════════════════════════════════════════════════
            self.workbook = Workbook()

            # Eliminar hoja por defecto
            if 'Sheet' in self.workbook.sheetnames:
                default_sheet = self.workbook['Sheet']
                self.workbook.remove(default_sheet)

            # ══════════════════════════════════════════════════════
            # PASO 2: Crear hoja de instrucciones (si se solicita)
            # ══════════════════════════════════════════════════════
            if self.config.get('options', {}).get('include_instructions', True):
                self._create_instructions_sheet()

            # ══════════════════════════════════════════════════════
            # PASO 3: Crear hoja de plantilla principal
            # ══════════════════════════════════════════════════════
            self._create_template_sheet()

            # ══════════════════════════════════════════════════════
            # PASO 4: Crear hoja de referencia de símbolos
            # ══════════════════════════════════════════════════════
            self._create_symbol_reference_sheet()

            # ══════════════════════════════════════════════════════
            # PASO 5: Guardar en memoria (BytesIO)
            # ══════════════════════════════════════════════════════
            excel_file = BytesIO()
            self.workbook.save(excel_file)
            excel_file.seek(0)

            logger.info(
                f"Plantilla generada exitosamente: "
                f"{len(self.workbook.sheetnames)} hojas creadas"
            )

            return excel_file

        except Exception as e:
            logger.error(f"Error al generar plantilla: {e}", exc_info=True)
            raise

    def _create_instructions_sheet(self):
        """
        Crea hoja de instrucciones con guía de uso.

        Hoja: "Instrucciones"
        Contenido:
        - Título del libro de trabajo
        - Instrucciones de uso
        - Información de auditoría
        - Notas importantes
        """
        ws = self.workbook.create_sheet("Instrucciones", 0)

        # Título
        ws.merge_cells('B2:F2')
        title_cell = ws['B2']
        title_cell.value = 'PLANTILLA DE MARCAS DE AUDITORÍA'
        title_cell.font = Font(bold=True, size=16, color=self.COLOR_HEADER_TEXT)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 30

        # Información del libro de trabajo
        ws.merge_cells('B4:F4')
        ws['B4'] = f"Libro de Trabajo: {self.config['workbook_name']}"
        ws['B4'].font = Font(bold=True, size=12)

        if 'audit_info' in self.config:
            ws.merge_cells('B5:F5')
            ws['B5'] = f"Auditoría: {self.config['audit_info'].get('title', 'N/A')}"
            ws['B5'].font = Font(size=11)

            ws.merge_cells('B6:F6')
            ws['B6'] = f"Creado por: {self.config['audit_info'].get('created_by', 'N/A')}"
            ws['B6'].font = Font(size=11)

        ws.merge_cells('B7:F7')
        ws['B7'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B7'].font = Font(size=11)

        # Instrucciones
        current_row = 9
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = 'INSTRUCCIONES DE USO:'
        ws[f'B{current_row}'].font = Font(bold=True, size=12, color=self.COLOR_HEADER_TEXT)
        ws[f'B{current_row}'].fill = PatternFill(
            start_color=self.COLOR_INSTRUCTIONS_BG,
            end_color=self.COLOR_INSTRUCTIONS_BG,
            fill_type='solid'
        )

        instructions = [
            "1. Diríjase a la hoja 'Plantilla' para ingresar los datos de auditoría.",
            "2. Complete cada columna según el tipo de datos especificado en el encabezado.",
            "3. Utilice los símbolos de auditoría de la hoja 'Referencia de Símbolos'.",
            "4. Las celdas marcadas con (*) son obligatorias.",
            "5. No modifique los encabezados de las columnas.",
            "6. Una vez completada, guarde el archivo y súbalo al sistema.",
            "7. El sistema validará el nombre del libro de trabajo automáticamente.",
        ]

        for idx, instruction in enumerate(instructions, start=1):
            current_row += 1
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = instruction
            ws[f'B{current_row}'].font = Font(size=10)
            ws[f'B{current_row}'].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[current_row].height = 25

        # Notas importantes
        current_row += 2
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = 'NOTAS IMPORTANTES:'
        ws[f'B{current_row}'].font = Font(bold=True, size=12, color='C00000')
        ws[f'B{current_row}'].fill = PatternFill(
            start_color='FFE699',
            end_color='FFE699',
            fill_type='solid'
        )

        notes = [
            f"• Este libro de trabajo está configurado para: {self.config['workbook_name']}",
            f"• Contiene {len(self.config['symbols'])} símbolos de auditoría disponibles",
            f"• Incluye {len(self.config['columns'])} columnas de datos configuradas",
            "• No elimine ni renombre las hojas de este archivo",
        ]

        for note in notes:
            current_row += 1
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = note
            ws[f'B{current_row}'].font = Font(size=10, italic=True)

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['G'].width = 2

        logger.debug("Hoja de instrucciones creada")

    def _create_template_sheet(self):
        """
        Crea la hoja de plantilla principal con columnas configuradas.

        Hoja: "Plantilla"
        Estructura:
        - Fila 1: Encabezados de columna
        - Filas 2-100: Espacio para datos (formato aplicado)
        """
        ws = self.workbook.create_sheet("Plantilla")

        # ═══════════════════════════════════════════════════════════
        # Crear encabezados de columna (Fila 1)
        # ═══════════════════════════════════════════════════════════
        border_style = Border(
            left=Side(style='thin', color=self.COLOR_DATA_BORDER),
            right=Side(style='thin', color=self.COLOR_DATA_BORDER),
            top=Side(style='thin', color=self.COLOR_DATA_BORDER),
            bottom=Side(style='thin', color=self.COLOR_DATA_BORDER)
        )

        for col_idx, column_config in enumerate(self.config['columns'], start=1):
            cell = ws.cell(row=1, column=col_idx)

            # Nombre de columna (agregar * si es obligatoria)
            column_name = column_config['name']
            if column_config.get('is_required', False):
                column_name += ' *'

            cell.value = column_name
            cell.font = Font(
                bold=True,
                size=11,
                color='FFFFFF'
            )
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER_TEXT,
                end_color=self.COLOR_HEADER_TEXT,
                fill_type='solid'
            )
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = border_style

            # Ajustar ancho de columna
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = column_config.get('width', 15)

        ws.row_dimensions[1].height = 30

        # ═══════════════════════════════════════════════════════════
        # Aplicar formato a filas de datos (Filas 2-100)
        # ═══════════════════════════════════════════════════════════
        data_rows = 100  # Número de filas predefinidas

        for row_idx in range(2, data_rows + 2):
            for col_idx, column_config in enumerate(self.config['columns'], start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Aplicar bordes
                cell.border = border_style

                # Aplicar formato según tipo de dato
                data_type = column_config.get('data_type', 'text')

                if data_type == 'number':
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                elif data_type == 'currency':
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif data_type == 'percentage':
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal='right')
                elif data_type == 'date':
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center')
                else:  # text
                    cell.alignment = Alignment(horizontal='left')

        # Congelar panel en la fila de encabezados
        ws.freeze_panes = 'A2'

        logger.debug(f"Hoja de plantilla creada con {len(self.config['columns'])} columnas")

    def _create_symbol_reference_sheet(self):
        """
        Crea hoja de referencia con todos los símbolos disponibles.

        Hoja: "Referencia de Símbolos"
        Columnas:
        - Símbolo
        - Descripción
        - Categoría
        """
        ws = self.workbook.create_sheet("Referencia de Símbolos")

        # Encabezados
        headers = ['Símbolo', 'Descripción', 'Categoría']
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER_TEXT,
                end_color=self.COLOR_HEADER_TEXT,
                fill_type='solid'
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style

        # Ajustar anchos
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15

        # Agregar símbolos
        color_code = self.config.get('options', {}).get('color_code_symbols', True)

        category_colors = {
            'verification': 'E2EFDA',  # Verde claro
            'calculation': 'FFF2CC',   # Amarillo claro
            'documentation': 'DEEBF7', # Azul claro
            'review': 'FCE4D6',        # Naranja claro
            'analysis': 'F4E4F4',      # Púrpura claro
            'custom': 'F2F2F2',        # Gris claro
        }

        for row_idx, symbol_data in enumerate(self.config['symbols'], start=2):
            # Símbolo
            symbol_cell = ws.cell(row=row_idx, column=1)
            symbol_cell.value = symbol_data['symbol']
            symbol_cell.font = Font(size=12, bold=True)
            symbol_cell.alignment = Alignment(horizontal='center', vertical='center')
            symbol_cell.border = border_style

            # Descripción
            desc_cell = ws.cell(row=row_idx, column=2)
            desc_cell.value = symbol_data['description']
            desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            desc_cell.border = border_style

            # Categoría
            cat_cell = ws.cell(row=row_idx, column=3)
            category = symbol_data.get('category', 'custom')
            cat_cell.value = category.capitalize()
            cat_cell.alignment = Alignment(horizontal='center', vertical='center')
            cat_cell.border = border_style

            # Aplicar color de categoría si está activado
            if color_code:
                bg_color = category_colors.get(category, 'FFFFFF')
                for col in [1, 2, 3]:
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color=bg_color,
                        end_color=bg_color,
                        fill_type='solid'
                    )

            ws.row_dimensions[row_idx].height = 25

        # Congelar panel
        ws.freeze_panes = 'A2'

        logger.debug(f"Hoja de referencia creada con {len(self.config['symbols'])} símbolos")

    def get_filename(self):
        """
        Genera el nombre de archivo para descarga.

        Returns:
            str: Nombre de archivo con timestamp
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        workbook_name = self.config['workbook_name'].replace(' ', '_')
        return f"{workbook_name}_Plantilla_{timestamp}.xlsx"
