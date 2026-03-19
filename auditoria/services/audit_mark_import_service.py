"""
Servicio para importar marcas de auditoría desde archivos Excel con filtrado por color.
"""

from openpyxl import load_workbook
from django.db import transaction
from django.core.exceptions import ValidationError
from auditoria.models import AuditMark
import logging

logger = logging.getLogger(__name__)


class AuditMarkImportService:
    """Servicio para importar marcas de auditoría desde Excel con filtrado de color."""

    GREEN_COLORS = [
        '00FF00', 'C6EFCE', '90EE90', '92D050', 'C5E0B4',
        '00B050', 'E2EFDA', 'A9D08E', '70AD47'
    ]
    YELLOW_COLORS = [
        'FFFF00', 'FFFFE0', 'FFF4CE', 'FFEB9C', 'FFD966',
        'FFC000', 'F4B084', 'FCE4D6'
    ]

    def __init__(self, audit_id, excel_file):
        self.audit_id = audit_id
        self.excel_file = excel_file
        self.marks_imported = 0
        self.marks_skipped_white = 0
        self.marks_skipped_yellow = 0
        self.marks_skipped_invalid = 0
        self.errors = []

    def validate_file(self):
        """Valida el archivo Excel cargado."""
        filename = self.excel_file.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            raise ValidationError("El archivo debe ser formato Excel (.xlsx o .xls)")

        if self.excel_file.size > 5 * 1024 * 1024:
            raise ValidationError("El archivo no debe superar 5MB")

        try:
            wb = load_workbook(self.excel_file, read_only=True)
            wb.close()
        except Exception as e:
            raise ValidationError(f"El archivo está corrupto o no se puede leer: {str(e)}")

    def _get_cell_background_color(self, cell):
        """Extrae el color de fondo de la celda."""
        if cell.fill and cell.fill.start_color:
            color = cell.fill.start_color
            if hasattr(color, 'rgb') and color.rgb:
                rgb = str(color.rgb)
                if len(rgb) == 8:
                    return rgb[2:].upper()
                return rgb.upper()
        return None

    def _is_green_row(self, row_cells):
        """Verifica si la fila tiene fondo verde."""
        for cell in row_cells:
            bg_color = self._get_cell_background_color(cell)
            if bg_color and any(green in bg_color for green in self.GREEN_COLORS):
                return True
        return False

    def _is_yellow_row(self, row_cells):
        """Verifica si la fila tiene fondo amarillo."""
        for cell in row_cells:
            bg_color = self._get_cell_background_color(cell)
            if bg_color and any(yellow in bg_color for yellow in self.YELLOW_COLORS):
                return True
        return False

    def parse_excel(self):
        """Analiza el archivo Excel y extrae las marcas."""
        wb = load_workbook(self.excel_file, data_only=True)
        ws = wb.active
        marks = []

        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                if self._is_yellow_row(row_cells):
                    self.marks_skipped_yellow += 1
                    continue

                if self._is_green_row(row_cells):
                    mark_data = self._extract_mark_from_row(row_cells, row_idx)
                    if mark_data:
                        marks.append(mark_data)
                    else:
                        self.marks_skipped_invalid += 1
                else:
                    has_data = any(cell.value for cell in row_cells[:4])
                    if has_data:
                        self.marks_skipped_white += 1

            except Exception as e:
                self.errors.append(f"Fila {row_idx}: Error - {str(e)}")
                self.marks_skipped_invalid += 1

        wb.close()
        return marks

    def _extract_mark_from_row(self, row_cells, row_idx):
        """Extrae datos de marca de una fila (3 columnas)."""
        symbol = str(row_cells[0].value).strip() if row_cells[0].value else None
        description = str(row_cells[1].value).strip() if row_cells[1].value else None
        work_paper = str(row_cells[2].value).strip() if len(row_cells) > 2 and row_cells[2].value else None

        if not symbol or not description:
            return None

        if description and ("Ejemplo:" in description or "Example:" in description):
            return None

        if work_paper and work_paper != work_paper.upper():
            self.errors.append(
                f"Fila {row_idx}: '{work_paper}' no está en MAYÚSCULAS. "
                f"Recomendado: '{work_paper.upper()}'"
            )

        return {
            'symbol': symbol,
            'description': description,
            'work_paper_number': work_paper
        }

    @transaction.atomic
    def import_marks(self, replace_existing=True):
        """Importa marcas a la base de datos."""
        self.validate_file()
        marks = self.parse_excel()

        if replace_existing:
            AuditMark.objects.filter(audit_id=self.audit_id).delete()

        if marks:
            mark_objects = [
                AuditMark(
                    audit_id=self.audit_id,
                    symbol=mark['symbol'],
                    description=mark['description'],
                    work_paper_number=mark['work_paper_number'],
                    is_active=True
                )
                for mark in marks
            ]
            AuditMark.objects.bulk_create(mark_objects)
            self.marks_imported = len(mark_objects)

        return {
            'success': True,
            'marks_imported': self.marks_imported,
            'marks_skipped_white': self.marks_skipped_white,
            'marks_skipped_yellow': self.marks_skipped_yellow,
            'marks_skipped_invalid': self.marks_skipped_invalid,
            'errors': self.errors
        }
